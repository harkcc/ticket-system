import os
import traceback
import openpyxl
import win32com.client
from abc import ABC, abstractmethod
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment,Color
import datetime
import re
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.image import Image
from logging.handlers import TimedRotatingFileHandler
import logging
import subprocess
import tkinter as tk
from tkinter import messagebox


# 日志
def mylogger(moudle_name, level=20,
             log_format="%(asctime)s,%(msecs)d %(name)s:%(levelname)s: [%(module)s(%(lineno)d)] %(message)s",
             date_format="%Y-%m-%d %H:%M:%S", console_print=True, log_file=None, when=None):
    # 防止产生多个logger，解决重复打印问题
    if moudle_name not in logging.Logger.manager.loggerDict:
        handle_flg = True
    else:
        handle_flg = False

    if console_print is False and log_file is None:
        print("Error: Save log in file, but input not log file!")
        return

    when = None if when not in ["S", "M", "H", "D", "midnight"] else when

    # create logger
    logger = logging.getLogger(moudle_name)
    logger.setLevel(level)
    formatter = logging.Formatter(log_format, date_format)
    if handle_flg is True:
        if console_print is True:
            # create handler，output log to console
            ch = logging.StreamHandler()
            ch.setFormatter(formatter)
            # logger add handler
            logger.addHandler(ch)
        if log_file:
            if when:
                logHandler = TimedRotatingFileHandler(log_file, when=when)
                logHandler.setFormatter(formatter)
                logger.addHandler(logHandler)
            else:
                fh = logging.FileHandler(log_file)
                fh.setFormatter(formatter)
                # logger add handler
                logger.addHandler(fh)
    return logger


'''全局变量'''
# 文件的地址
message_file = None
product_file = None
product_image = None
message_folder = os.path.join(os.getcwd(), "相关信息")
current_directory = os.getcwd()  # 获取当前工作目录
# wuliu_list[] = ['德邦空派','叮铛美森快船','林道UPS红单','林道海运带电带磁','叮铛卡铁带电带磁','叮铛海运带电带磁','德邦美森','递信日本空派','叮铛卡航','USP红单DJ票','顺丰空派','USP红单KU票']
flag_battery = False   #文件中是否有带电带磁的选项


# 设置要创建文件夹的路径和名称
log_directory = os.path.join(current_directory, 'log')
output_directory = os.path.join(current_directory, 'output')
history_directory = os.path.join(current_directory, 'history')

# 创建文件夹
os.makedirs(log_directory, exist_ok=True)
os.makedirs(output_directory, exist_ok=True)
os.makedirs(history_directory, exist_ok=True)

#日志文件
log_file = os.path.join(log_directory, 'log.log')
logger_file = os.path.join(log_directory, 'loghistory.log')
log = mylogger("log", console_print=True, level=20, log_file=log_file)
#logger = mylogger('history_record', log_file=logger_file)


# 存放信息的数组
account = []  # 账户信息
image_data = [] # 存放带有图片的sku的文件名

# 装数据的字典
product_dict = {}
box_rule_dict = {}
box_dict = {}
relate_folder = os.path.join(os.getcwd(), "相关信息")


'''首先是查找该有文件的位置'''
# 查找相关文件，记录文件的位置

for f in os.listdir(message_folder):
    if '箱子账号信息' in f:
        message_file = os.path.join(message_folder, f)
    # else:
    #     log.error("箱子账号信息文件不存在或者文件名错误")
    if '产品导出' in f:
        product_file = os.path.join(message_folder, f)
    # else:
    #     log.error("产品导出信息不存在或者文件名错误")
    if '产品图片' in f:
        product_image = os.path.join(message_folder, f)
    # else:
    #     log.error("产品图片信息不存在或者文件名错误")

# 要给为日志错误的形式，用raise
# 这里的检查可以统一成一个方法
if message_file != None:
    wb_message = openpyxl.load_workbook(message_file)
    sheet_box = wb_message['箱子命名规则']
    sheet_account = wb_message['账号信息']
else:
    log.error("文件缺失")

if product_file != None:
    wb_product = openpyxl.load_workbook(product_file)
    sheet_product = wb_product['Sheet1']
else:
    log.error("文件缺失")

if product_image != None:
    for filename in os.listdir(product_image):
        #这里可能要改？
        image_data.append(filename)
else:
    log.error("文件缺失")


'''搭建相应的类'''
# 数据类的建立：箱子的规定信息-box_rule，产品的信息:product 实际箱子的信息：box_translation
class box_rule:
    def __init__(self, box_type, length, width, height, box_weight):
        """
        箱子类，包含以下属性：
        box_type: str, 箱子类型
        weight: float, 箱子重量
        length: float, 箱子长度
        width: float, 箱子宽度
        height: float, 箱子高度
        """
        self.box_type = box_type
        self.length = length
        self.width = width
        self.height = height
        self.box_weight = box_weight

# 商品信息类
class Product:
    def __init__(self, sku, cn_name, en_name, price, cn_material, en_material, cn_usage, en_usage, model, hs_code, link,
                 has_electricity, has_magnetism, brand, weight, asin, list_fee, out_fee, page, quantity,box_quantity):
        """
        产品类，包含以下属性：
        sku: str, 产品SKU
        cn_name: str, 产品中文名
        en_name: str, 产品英文名
        price: float, 产品价格
        cn_material: str, 产品中文材质
        en_material: str, 产品英文材质
        cn_usage: str, 产品中文用途
        en_usage: str, 产品英文用途
        model: str, 产品型号
        hs_code: str, 产品海关编码
        link: str, 产品链接
        has_electricity: bool, 产品是否带电
        has_magnetism: bool, 产品是否带磁
        brand: str, 产品品牌
        weight: float, 产品重量
        asin: str, 产品ASIN
        list_fee: float, 产品上架费用
        out_fee: float, 产品出库费用
        quantity: int 总数量
        box_quantity:存放在对应箱子的数量
        """
        self.sku = sku
        self.cn_name = cn_name
        self.en_name = en_name
        self.price = price
        self.cn_material = cn_material
        self.en_material = en_material
        self.cn_usage = cn_usage
        self.en_usage = en_usage
        self.model = model
        self.hs_code = hs_code
        self.link = link
        self.has_electricity = has_electricity
        self.has_magnetism = has_magnetism
        self.brand = brand
        self.weight = weight
        self.asin = asin
        self.list_fee = list_fee
        self.out_fee = out_fee
        self.page = page
        self.quantity = quantity
        self.box_quantity = box_quantity

# 实际装有商品的箱子类
class box_translation:
    def __init__(self, box_number, box_message, weight, array=None):
        self.box_number = box_number
        self.box_message = box_message
        self.array = array if array is not None else []
        self.weight = weight

# 专用于日本宏川
class box_identity:
    def __init__(self,box_num,inner_self, inner_FN,product_id,box_quantity,inner_product):
        """
        inner_self :自编内标
        inner_FN: 亚马逊标
        box_num:箱号
        prodcut_id：产品标题
        box_quantity:数量
        inner_prodecut:自编货件号
        """
        self.inner_self = inner_self
        self.inner_FN = inner_FN
        self.box_num = box_num
        self.product_id = product_id
        self.box_quantity = box_quantity
        self.inner_product = inner_product

# 获取箱子的规定信息
for row in sheet_box.iter_rows(min_row=2, values_only=True):
    # 从第一列获取箱子编号
    box_type = row[0]
    # 从第二列获取重量
    weight = row[4]
    # 从第三列获取长度
    length = row[1]
    # 从第四列获取宽度
    width = row[2]
    # 从第五列获取高度
    height = row[3]
    # 创建Box对象并将其添加到字典中
    box_rule_dict[box_type] = box_rule(box_type, length, width, height, weight)



# 工厂类-用于模板
class InvoiceTemplateFactory:
    def create_template(self, logistics_provider):
        if logistics_provider == "A":   # 百泰FBC858672-2023.03.02-林道海运带电带磁QT票-4件-法国-发票装箱单
            return CDMS_franch_lin()
        elif logistics_provider == "B": # 百泰FBA15GL7CSKH-2023.01.09-叮铛卡航限时达QN票-6件-德国-发票装箱单.xlsx
            return dingdan_germany()
        elif logistics_provider == "C": # 百泰FBA15GQLWK0L-2023.03.03-顺丰空派QA票-3件-英国-发票装箱单.xlsx
            return sf_eg()
        elif logistics_provider == "D": # 百泰FBC851061-2023.01.12-叮铛卡铁带磁带电QB票-5件-法国-发票装箱单.xls
            return dingdangkatie_franch()
        elif logistics_provider == "E": # 百泰FBC844244-2022.12.13-叮铛海运带磁带电QL票-2件-法国-发票装箱单.xlsx
            return dingdanghaiyun_franch()
        elif logistics_provider == "F": #百泰FBA16WW8DBN0-2022.09.28-叮铛美森快船统配DL票-5件-美国-发票装箱单.xlsx
            return dingdangmeisheng_usa()
        elif logistics_provider == "G":
            return debangmeisheng_usa()
        elif logistics_provider == "H":
            return shengzyiruoda_new()
        elif logistics_provider == "I":
            return lindaoups_usa()
        elif logistics_provider == "J":
            return hongdangDJ_usa()
        elif logistics_provider == "K":
            return gemDG_usa()
        elif logistics_provider == "L":
            return hongdangKU_eng()
        elif logistics_provider == "M":
            return dixing_jap()
        elif logistics_provider == "N":
            return hongchuang_jap()
        elif logistics_provider == "O":
            return lindao_jap()

        else:
            raise ValueError(f"Unsupported logistics provider: {logistics_provider}")

'''图片处理'''

# 插入图片的方法
def insert_centered_image(worksheet, cell_address, image_path, fixed_width=None, fixed_height=None):
    # 读取图片
    img = Image(image_path)

    # 获取列字母和行号
    col_letter = cell_address[0]
    row_num = int(cell_address[1:])

    # 计算单元格宽度和高度的像素值
    if fixed_width is not None and fixed_height is not None:
        cell_width_px = fixed_width
        cell_height_px = fixed_height
    else:
        column_dimensions = worksheet.column_dimensions
        column_width = column_dimensions[col_letter].width
        cell_width = (column_width - 1) / 7 * 140  # 7 pixels per character, 140 pixels per cell
        row_dimensions = worksheet.row_dimensions
        row_height = row_dimensions[row_num].height
        cell_height = (row_height - 1) / 15 * 20  # 15 pixels per point, 20 points per cell
        cell_width_px = cell_width
        cell_height_px = cell_height

    # 获取图片的宽度和高度
    width, height = img.width, img.height

    # 计算缩放比例
    scale_width = cell_width_px / width
    scale_height = cell_height_px / height
    scale = min(scale_width, scale_height)

    # 缩放图片
    img.width, img.height = width * scale, height * scale

    # 将图片插入到目标单元格
    img.anchor = cell_address
    worksheet.add_image(img)

# 根据相关的位置信息，向表格中插入图片
def insert_images(cell_address,col_name,sheet, new_folder_path, row_num, sku,height):
    global image_data
    if '/' in sku:
        sku = sku.replace("/", '=')
    for index in image_data:
        if sku in index:
            column_dimensions = sheet.column_dimensions
            column_width = column_dimensions[cell_address[0]].width
            cell_width = (column_width - 1) / 7 * 140  # 7 pixels per character, 140 pixels per cell

            row_height = height

            cell_height = (row_height - 1) / 17 * 20  # 15 pixels per point, 20 points per cell
            fixed_width = cell_width
            fixed_height = cell_height

            path = os.path.join(new_folder_path, index)
            cell_address = f"{col_name}{row_num}"
            insert_centered_image(sheet, cell_address, path,fixed_width=fixed_width, fixed_height=fixed_height)

# 发票模版抽象类
class InvoiceTemplate(ABC):
    @abstractmethod
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        pass


''' 生成发票输出的名字'''
#专门生成发票模板的名字
def creat_output_name(folder_name):

    part = folder_name[0].split("-")
    pattern = r'\d{4}\.\d{2}\.\d{2}'  # 匹配 yyyy.mm.dd 格式的日期
    folder_name =  re.sub(pattern, '', folder_name[0])
    # 这里分很多种情况，判断前一个或者两个符不符合命名条件这些

    mid_account = part[0] + '-' + part[1]
    max_account = part[0] + '-' + part[1] + '-' + part[2]
    if len(part) < 2:
        # 处理 part 列表不至少有两个元素的情况
        log.error("无效的文件名格式，请检查格式是装箱模板 业务账号（例如：CDMS-发票号(FBA*****)....")
        return None

    if  max_account in account:
        folder_name = folder_name.replace(max_account+'-', '', 1)   #摘去原来文件中的时间
        result = '-'+(folder_name.split('-')[1:])[0] + '-' + (folder_name.split('-')[1:])[1] + '-' + \
                 (folder_name.split('-')[1:])[2][3:-2]
        return result
    elif mid_account in account:
        folder_name = folder_name.replace(mid_account+'-', '', 1)
        result = '-'+(folder_name.split('-')[1:])[0]+'-'+(folder_name.split('-')[1:])[1]+'-'+(folder_name.split('-')[1:])[2][3:-2]
        return result
    elif part[0] in account:
        folder_name = folder_name.replace(part[0] + '-', '', 1)
        result = '-'+(folder_name.split('-')[1:])[0] + '-' + (folder_name.split('-')[1:])[1] + '-' + \
                 (folder_name.split('-')[1:])[2][3:-2]
        return result
    else:
        log.error("业务账号有误，请检查这个业务号是不是不在相关信息中账号信息的表格中")



'''欧美类的模板，主要是格式细节差异'''

# 发票模版 A-百泰FBC858672-2023.03.02-林道海运带电带磁QT票-4件-法国-发票装箱单.xlsx
class CDMS_franch_lin(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        # 打开对应的模板文件
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['模板']
        # 记录第17行的行高列宽，17行第一列的颜色和边框格式，和第15列的边框格式
        row_height = sheet.row_dimensions[17].height
        cell_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))
        cell_Alignment = Alignment(horizontal='center', vertical='center')
        cell_fill = sheet.cell(row=17, column=1).fill
        cell_style = sheet.cell(row=17, column=1)
        cell_font = Font(name=cell_style.font.name, size=cell_style.font.size, bold=cell_style.font.bold,
                         italic=cell_style.font.italic,
                         vertAlign=cell_style.font.vertAlign, color=cell_style.font.color)

        #消除合并单元格
        for ranges in ['B17:B28', 'C17:C28', 'D17:D28', 'E17:E28',
                      'B29:B41', 'C29:C41', 'D29:D41', 'E29:E41',
                      'B42:B54', 'C42:C54', 'D42:D54', 'E42:E54',
                      'B55:B63', 'C55:C63', 'D55:D63', 'E55:E63','F13:H13','F14:H14','F1:H1','F2:H2']:
            sheet.unmerge_cells(ranges)

        # 删除所有图片的方法，
        sheet._images = []

        # 先将所有的行高变为普通的，避免少于原表格的情况
        for r in range(17, 64):
            sheet.row_dimensions[r].height = sheet.row_dimensions[70].height

        #删除原来表格中的内容
        sheet.delete_rows(17, sheet.max_row - 15)

        #新建一个临时文件，在临时文件中操作
        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)


        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket['模板']

        #遍历数据，写入表格
        row_num = 17
        for i, (box_number, box) in enumerate(dict_box.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = dict_product[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    new_sheet.cell(row=row_num, column=1, value=box_number)
                    new_sheet.cell(row=row_num, column=2, value=box.weight)
                    new_sheet.cell(row=row_num, column=3, value=box.box_message.length)
                    new_sheet.cell(row=row_num, column=4, value=box.box_message.width)
                    new_sheet.cell(row=row_num, column=5, value=box.box_message.height)
                    new_sheet.cell(row=row_num, column=7, value=product.cn_name)
                    new_sheet.cell(row=row_num, column=6, value=product.en_name)
                    new_sheet.cell(row=row_num, column=8, value=product.price)
                    new_sheet.cell(row=row_num, column=10, value=product.en_material)
                    new_sheet.cell(row=row_num, column=12, value=product.en_usage)
                    new_sheet.cell(row=row_num, column=14, value=product.model)
                    new_sheet.cell(row=row_num, column=11, value=product.hs_code)
                    new_sheet.cell(row=row_num, column=17, value=product.link)
                    new_sheet.cell(row=row_num, column=13, value=product.brand)
                    new_sheet.cell(row=row_num, column=18, value=product.weight)
                    new_sheet.cell(row=row_num, column=9, value=product.box_quantity[box_number])
                    new_sheet.cell(row=row_num, column=16, value=product.box_quantity[box_number] * product.price)
                    new_sheet.cell(row=row_num, column=20, value=sku)
                    insert_images('O17','O',new_sheet,product_image,row_num,sku,row_height)
                    row_num += 1
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=2, end_row=row_num - 1, end_column=2)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=3, end_row=row_num - 1, end_column=3)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=4, end_row=row_num - 1, end_column=4)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=5, end_row=row_num - 1, end_column=5)

        data_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        #调整格式，字体，边框....
        for row in new_sheet.iter_rows(min_row=16, min_col=1, max_col=20, values_only=False):
            for cell in row:
                cell.border = data_border
                cell.font = cell_font
                cell.alignment = cell_Alignment

        new_sheet.cell(row=16, column=20, value="产品SKU")

        # 发票号填写
        new_sheet.cell(row=13, column=6, value=ticket_name[1])
        new_sheet.cell(row=14, column=6, value=ticket_name[1])
        new_sheet.merge_cells(start_row=13, start_column=6, end_row=13,
                              end_column=8)
        new_sheet.merge_cells(start_row=14, start_column=6, end_row=14,
                              end_column=8)


        #带点带磁填写
        if '电' in ticket_name[2][0]:
            new_sheet.cell(row=1, column=6, value='是')
        else:
            new_sheet.cell(row=1, column=6, value='否')
        new_sheet.merge_cells(start_row=1, start_column=6, end_row=1,
                              end_column=8)

        if '磁' in ticket_name[2][0]:
            new_sheet.cell(row=2, column=6, value='是')
        else:
            new_sheet.cell(row=2, column=6, value='否')
        new_sheet.merge_cells(start_row=2, start_column=6, end_row=2,
                              end_column=8)

        # 边框设置
        for row in new_sheet.iter_rows(min_row=17, min_col=17, max_col=17, values_only=False):
            for cell in row:
                cell.border = cell_border

        # 颜色模型配置
        fill = PatternFill(start_color=cell_fill.start_color, end_color=cell_fill.end_color,
                           fill_type=cell_fill.fill_type)

        # 1和17行的颜色设置
        for row in new_sheet.iter_rows(min_row=17, min_col=1, max_col=1):
            for cell in row:
                cell.fill = fill
        for row in new_sheet.iter_rows(min_row=17, min_col=17, max_col=17):
            for cell in row:
                cell.fill = fill
        # 多出来
        for row in new_sheet.iter_rows(min_row=17, values_only=False):
            for cell in row:
                new_sheet.row_dimensions[cell.row].height = row_height

        # 这里的名字还需要再改动一下
        # wb.save(os.path.join(os.getcwd(), "output", model_file_name))
        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 发票模版 B -百泰FBA15GL7CSKH-2023.01.09-叮铛卡航限时达QN票-6件-德国-发票装箱单.xlsx
class dingdan_germany(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):

        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['模板']

        # 删除所有图片的方法，
        sheet._images = []

        row_height = sheet.row_dimensions[18].height

        for r in range(18, 29):
            sheet.row_dimensions[r].height = sheet.row_dimensions[40].height

        cell_style = sheet.cell(row=18, column=1)
        cell_font = Font(name=cell_style.font.name, size=cell_style.font.size, bold=cell_style.font.bold,
                         italic=cell_style.font.italic,
                         vertAlign=cell_style.font.vertAlign, color=cell_style.font.color)
        cell_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))
        cell_Alignment = Alignment(horizontal='center', vertical='center')

        # 删除原来的内容
        sheet.delete_rows(18, 29)

        number_ticket = ticket_name[1]+'U00000'
        row_num = 18
        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    FBA_number = number_ticket + str(box_number)
                    sheet.cell(row=row_num, column=1, value=FBA_number)
                    sheet.cell(row=row_num, column=2, value=box.weight)
                    sheet.cell(row=row_num, column=3, value=product.en_name)
                    sheet.cell(row=row_num, column=4, value=product.cn_name)
                    sheet.cell(row=row_num, column=5, value=product.price)
                    sheet.cell(row=row_num, column=6, value=product.box_quantity[box_number])
                    sheet.cell(row=row_num, column=7, value=product.en_material)
                    sheet.cell(row=row_num, column=8, value=product.hs_code)
                    usage = product.cn_usage + '' + product.en_usage
                    sheet.cell(row=row_num, column=9, value=usage)
                    sheet.cell(row=row_num, column=10, value=product.brand)
                    sheet.cell(row=row_num, column=11, value=product.model)
                    sheet.cell(row=row_num, column=12, value=product.link)
                    sheet.cell(row=row_num, column=15, value=product.quantity * product.price)
                    # 这里有疑问
                    sheet.cell(row=row_num, column=16, value=box.weight)
                    sheet.cell(row=row_num, column=17, value=box.box_message.length)
                    sheet.cell(row=row_num, column=18, value=box.box_message.width)
                    sheet.cell(row=row_num, column=19, value=box.box_message.height)
                    sheet.cell(row=row_num, column=20, value=sku)
                    insert_images('N18', 'N', sheet, product_image, row_num, sku, row_height)
                    row_num += 1

        data_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        for row in sheet.iter_rows(min_row=18, min_col=1, max_col=20, values_only=False):
            for cell in row:
                cell.border = data_border
                cell.font = cell_font
                cell.alignment = cell_Alignment

        # 还有一些细节改一下就好
        for row in sheet.iter_rows(min_row=18, min_col=21, max_col=21, values_only=False):
            for cell in row:
                cell.border = cell_border

        # 行高
        for row in sheet.iter_rows(min_row=18, values_only=False):
            for cell in row:
                sheet.row_dimensions[cell.row].height = row_height


        # 修改表头信息
        for ranges in ['B1:D1', 'F1:G1','B16:D16']:
            sheet.unmerge_cells(ranges)

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket['模板']

        # 带点带磁填写
        if '电' in ticket_name[2][0]:
            new_sheet.cell(row=1, column=6, value='是')
        else:
            new_sheet.cell(row=1, column=6, value='否')
        new_sheet.merge_cells(start_row=1, start_column=6, end_row=1,
                              end_column=7)

        if '磁' in ticket_name[2][0]:
            new_sheet.cell(row=2, column=6, value='是')
        else:
            new_sheet.cell(row=2, column=6, value='否')
        new_sheet.merge_cells(start_row=2, start_column=6, end_row=2,
                              end_column=7)

        new_sheet.cell(row=1, column=2, value=ticket_name[1])
        new_sheet.merge_cells(start_row=1, start_column=2, end_row=1,
                              end_column=4)

        new_sheet.cell(row=16, column=2, value=len(box_dict))
        new_sheet.merge_cells(start_row=16, start_column=2, end_row=16,
                              end_column=4)

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 模板C--百泰FBA15GQLWK0L-2023.03.03-顺丰空派QA票-3件-英国-发票装箱单.xlsx
class sf_eg(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['Sheet1']

        # 删除所有图片的方法
        sheet._images = []

        row_height = sheet.row_dimensions[12].height

        cell_page = sheet.cell(row=12, column=12)
        cell_border = Border(left=cell_page.border.left, right=cell_page.border.right, top=cell_page.border.top,
                             bottom=cell_page.border.bottom)
        cell_style = sheet.cell(row=12, column=1)
        cell_font = Font(name=cell_style.font.name, size=cell_style.font.size, bold=cell_style.font.bold,
                         italic=cell_style.font.italic,
                         vertAlign=cell_style.font.vertAlign, color=cell_style.font.color)
        cell_Alignment = Alignment(horizontal='center', vertical='center')

        for r in range(12, 15):
            sheet.row_dimensions[r].height = sheet.row_dimensions[20].height

        # 删除原来的内容
        sheet.delete_rows(12, 15)

        number_ticket = ticket_name[1]+'U00000'
        row_num = 12
        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    FBA_number = number_ticket + str(box_number)
                    sheet.cell(row=row_num, column=1, value=FBA_number)
                    sheet.cell(row=row_num, column=2, value=sku)
                    sheet.cell(row=row_num, column=3, value=product.en_name)
                    sheet.cell(row=row_num, column=4, value=product.cn_name)
                    sheet.cell(row=row_num, column=5, value=product.brand)
                    sheet.cell(row=row_num, column=6, value=product.model)
                    sheet.cell(row=row_num, column=7, value=product.cn_material)
                    sheet.cell(row=row_num, column=8, value=product.en_material)
                    usage = product.cn_usage + '' + product.en_usage
                    sheet.cell(row=row_num, column=9, value=usage)
                    sheet.cell(row=row_num, column=10, value='包装')
                    sheet.cell(row=row_num, column=11, value=product.hs_code)
                    sheet.cell(row=row_num, column=12, value=product.box_quantity[box_number])
                    sheet.cell(row=row_num, column=13, value=product.price)
                    sheet.cell(row=row_num, column=14, value=product.box_quantity[box_number] * product.price)
                    sheet.cell(row=row_num, column=15, value=box.box_message.length)
                    sheet.cell(row=row_num, column=16, value=box.box_message.width)
                    sheet.cell(row=row_num, column=17, value=box.box_message.height)
                    sheet.cell(row=row_num, column=18, value=box.weight)
                    sheet.cell(row=row_num, column=19, value=product.link)
                    sheet.cell(row=row_num, column=20, value='')
                    insert_images('T12', 'T', sheet, product_image, row_num, sku, row_height)
                    row_num += 1

        data_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        for row in sheet.iter_rows(min_row=12, min_col=1, max_col=20, values_only=False):
            for cell in row:
                cell.border = data_border
                cell.font = cell_font
                cell.alignment = cell_Alignment

        col_width = sheet.column_dimensions['S'].width
        for row in sheet.iter_rows(min_row=12, min_col=12):
            for cell in row:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                sheet.column_dimensions[col_letter].width = col_width

        for row in sheet.iter_rows(min_row=12, min_col=19, max_col=19, values_only=False):
            for cell in row:
                cell.border = cell_border

        # 多出来
        for row in sheet.iter_rows(min_row=12, values_only=False):
            for cell in row:
                sheet.row_dimensions[cell.row].height = row_height



        for ranges in ['M7:T7']:
            sheet.unmerge_cells(ranges)

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket['Sheet1']


        new_sheet.cell(row=7, column=13, value=ticket_name[1])
        new_sheet.merge_cells(start_row=7, start_column=13, end_row=7,
                              end_column=20)


        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 模板D，百泰FBC851061-2023.01.12-叮铛卡铁带磁带电QB票-5件-法国-发票装箱单.xls
class dingdangkatie_franch(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['Sheet1']

        # 删除所有图片的方法
        sheet._images = []

        row_height = sheet.row_dimensions[18].height

        # 记录第17行的行高列宽，17行第一列的颜色和边框格式，和第15列的边框格式

        cell_page = sheet.cell(row=18, column=1)
        cell_border = Border(left=cell_page.border.left, right=cell_page.border.right, top=cell_page.border.top,
                             bottom=cell_page.border.bottom)
        cell_style = sheet.cell(row=18, column=1)
        cell_font = Font(name=cell_style.font.name, size=cell_style.font.size, bold=cell_style.font.bold,
                         italic=cell_style.font.italic,
                         vertAlign=cell_style.font.vertAlign, color=cell_style.font.color)
        cell_Alignment = Alignment(horizontal='center', vertical='center')

        for r in range(18, 51):
            sheet.row_dimensions[r].height = sheet.row_dimensions[60].height

        # 删除原来的内容
        sheet.delete_rows(18, 51)

        row_num = 18
        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    sheet.cell(row=row_num, column=1, value=box_number)
                    sheet.cell(row=row_num, column=2, value=box.weight)
                    sheet.cell(row=row_num, column=3, value=box.box_message.length)
                    sheet.cell(row=row_num, column=4, value=box.box_message.width)
                    sheet.cell(row=row_num, column=5, value=box.box_message.height)
                    sheet.cell(row=row_num, column=6, value=product.en_name)
                    sheet.cell(row=row_num, column=7, value=product.cn_name)
                    sheet.cell(row=row_num, column=8, value=product.price)
                    sheet.cell(row=row_num, column=9, value=product.box_quantity[box_number])
                    sheet.cell(row=row_num, column=10, value='件')
                    sheet.cell(row=row_num, column=11, value=product.en_material)
                    sheet.cell(row=row_num, column=12, value=product.hs_code)
                    usage = product.cn_usage + '' + product.en_usage
                    sheet.cell(row=row_num, column=13, value=usage)
                    sheet.cell(row=row_num, column=14, value=product.brand)
                    sheet.cell(row=row_num, column=15, value=product.model)
                    # 型号是什么
                    sheet.cell(row=row_num, column=17, value=product.link)

                    sheet.cell(row=row_num, column=23, value=sku)
                    insert_images('R18', 'R', sheet, product_image, row_num, sku, row_height)
                    row_num += 1

        data_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        for row in sheet.iter_rows(min_row=18, min_col=1, max_col=23, values_only=False):
            for cell in row:
                cell.border = data_border
                cell.font = cell_font
                cell.alignment = cell_Alignment

        for row in sheet.iter_rows(min_row=18, min_col=23, max_col=23, values_only=False):
            for cell in row:
                cell.border = cell_border

        # 多出来
        for row in sheet.iter_rows(min_row=18, values_only=False):
            for cell in row:
                sheet.row_dimensions[cell.row].height = row_height


        for ranges in ['B2:D2','F2:H2','F3:H3','B16:D16']:
            sheet.unmerge_cells(ranges)

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket['Sheet1']

        if '电' in ticket_name[2][0]:
            new_sheet.cell(row=2, column=6, value='是')
        else:
            new_sheet.cell(row=2, column=6, value='否')
        new_sheet.merge_cells(start_row=2, start_column=6, end_row=2,
                              end_column=8)

        if '磁' in ticket_name[2][0]:
            new_sheet.cell(row=3, column=6, value='是')
        else:
            new_sheet.cell(row=3, column=6, value='否')
        new_sheet.merge_cells(start_row=3, start_column=6, end_row=3,
                              end_column=8)

        new_sheet.cell(row=2, column=2, value=ticket_name[1])
        new_sheet.merge_cells(start_row=2, start_column=2, end_row=2,
                              end_column=4)

        new_sheet.cell(row=16, column=2, value=len(box_dict))
        new_sheet.merge_cells(start_row=16, start_column=2, end_row=16,
                              end_column=4)

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

#模板E 百泰FBC844244-2022.12.13-叮铛海运带磁带电QL票-2件-法国-发票装箱单.xlsx
class dingdanghaiyun_franch(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['模板']

        # 删除所有图片的方法
        sheet._images = []

        row_height = sheet.row_dimensions[28].height
        cell_page = sheet.cell(row=28, column=1)
        cell_border = Border(left=cell_page.border.left, right=cell_page.border.right, top=cell_page.border.top,
                             bottom=cell_page.border.bottom)
        cell_style = sheet.cell(row=28, column=1)
        cell_font = Font(name=cell_style.font.name, size=cell_style.font.size, bold=cell_style.font.bold,
                         italic=cell_style.font.italic,
                         vertAlign=cell_style.font.vertAlign, color=cell_style.font.color)
        cell_Alignment = Alignment(horizontal='center', vertical='center')

        row_num = 28

        for r in range(28, 56):
            sheet.row_dimensions[r].height = sheet.row_dimensions[60].height

        # 删除原来的内容
        sheet.delete_rows(28, 56)

        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    sheet.cell(row=row_num, column=1, value=box_number)
                    sheet.cell(row=row_num, column=2, value=box.weight)
                    sheet.cell(row=row_num, column=3, value=box.box_message.length)
                    sheet.cell(row=row_num, column=4, value=box.box_message.width)
                    sheet.cell(row=row_num, column=5, value=box.box_message.height)
                    sheet.cell(row=row_num, column=6, value=product.en_name)
                    sheet.cell(row=row_num, column=7, value=product.cn_name)
                    sheet.cell(row=row_num, column=8, value=product.price)
                    sheet.cell(row=row_num, column=9, value=product.box_quantity[box_number])
                    material = product.en_material + '' + product.cn_material
                    sheet.cell(row=row_num, column=10, value=material)
                    sheet.cell(row=row_num, column=11, value=product.hs_code)
                    usage = product.cn_usage + '' + product.en_usage
                    sheet.cell(row=row_num, column=12, value=usage)
                    # 给的模板有一点问题
                    sheet.cell(row=row_num, column=13, value=product.brand)
                    sheet.cell(row=row_num, column=14, value=product.model)
                    sheet.cell(row=row_num, column=15, value=product.link)
                    insert_images('Q28', 'Q', sheet, product_image, row_num, sku, row_height)
                    sheet.cell(row=row_num, column=21, value=sku)
                    row_num += 1

        data_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        for row in sheet.iter_rows(min_row=28, min_col=1, max_col=21, values_only=False):
            for cell in row:
                cell.border = data_border
                cell.font = cell_font
                cell.alignment = cell_Alignment

        for row in sheet.iter_rows(min_row=28, min_col=16, max_col=21, values_only=False):
            for cell in row:
                cell.border = cell_border

        # 多出来
        for row in sheet.iter_rows(min_row=28, values_only=False):
            for cell in row:
                sheet.row_dimensions[cell.row].height = row_height


        for ranges in ['B14:D14','B15:D15','B16:D16','B17:D17']:
            sheet.unmerge_cells(ranges)

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket['模板']

        if '电' in ticket_name[2][0]:
            new_sheet.cell(row=16, column=2, value='是')
        else:
            new_sheet.cell(row=16, column=2, value='否')
        new_sheet.merge_cells(start_row=16, start_column=2, end_row=16,
                              end_column=4)

        if '磁' in ticket_name[2][0]:
            new_sheet.cell(row=17, column=2, value='是')
        else:
            new_sheet.cell(row=17, column=2, value='否')
        new_sheet.merge_cells(start_row=17, start_column=2, end_row=17,
                              end_column=4)

        new_sheet.cell(row=14, column=2, value=ticket_name[1])
        new_sheet.merge_cells(start_row=14, start_column=2, end_row=14,
                              end_column=4)

        new_sheet.cell(row=15, column=2, value=len(box_dict))
        new_sheet.merge_cells(start_row=15, start_column=2, end_row=15,
                              end_column=4)

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 模板F -百泰FBA16WW8DBN0-2022.09.28-叮铛美森快船统配DL票-5件-美国-发票装箱单.xlsx
class dingdangmeisheng_usa(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['清关发票']

        # 删除所有图片的方法
        sheet._images = []

        row_height = sheet.row_dimensions[12].height
        cell_page = sheet.cell(row=12, column=1)
        cell_border = Border(left=cell_page.border.left, right=cell_page.border.right, top=cell_page.border.top,
                             bottom=cell_page.border.bottom)
        cell_style = sheet.cell(row=12, column=1)
        cell_font = Font(name=cell_style.font.name, size=cell_style.font.size, bold=cell_style.font.bold,
                         italic=cell_style.font.italic,
                         vertAlign=cell_style.font.vertAlign, color=cell_style.font.color)
        cell_Alignment = Alignment(horizontal='center', vertical='center')

        for r in range(12, 17):
            sheet.row_dimensions[r].height = sheet.row_dimensions[20].height

        # 删除原来的内容
        sheet.delete_rows(12, 17)

        row_num = 12
        number_ticket = ticket_name[1] + 'U00000'
        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    FBA_number = number_ticket + str(box_number)
                    sheet.cell(row=row_num, column=1, value=FBA_number)
                    # 这里有一个Reference ID必填？？？
                    # 这里由于限制先不填sku
                    volume = str(box.box_message.length) + '*' + str(box.box_message.width) + '*' + str(
                        box.box_message.height)
                    sheet.cell(row=row_num, column=3, value=volume)
                    sheet.cell(row=row_num, column=4, value=box_number)
                    # 这里是毛重----后面要商量一下
                    sheet.cell(row=row_num, column=5, value=box.weight)
                    sheet.cell(row=row_num, column=6, value=box.weight)

                    sheet.cell(row=row_num, column=7, value=product.en_name)
                    sheet.cell(row=row_num, column=8, value=product.cn_name)

                    sheet.cell(row=row_num, column=9, value=product.box_quantity[box_number])
                    sheet.cell(row=row_num, column=10, value=product.price)
                    sheet.cell(row=row_num, column=11, value=product.brand)
                    sheet.cell(row=row_num, column=12, value=product.model)
                    sheet.cell(row=row_num, column=13, value=product.cn_material)
                    usage = product.cn_usage + '' + product.en_usage
                    sheet.cell(row=row_num, column=14, value=usage)
                    insert_images('O12', 'O', sheet, product_image, row_num, sku, row_height)
                    row_num += 1

        sheet.cell(row=3, column=6, value=len(box_dict))

        data_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        for row in sheet.iter_rows(min_row=12, min_col=1, max_col=15, values_only=False):
            for cell in row:
                cell.border = data_border
                cell.font = cell_font
                cell.alignment = cell_Alignment

        for row in sheet.iter_rows(min_row=12, min_col=15, max_col=15, values_only=False):
            for cell in row:
                cell.border = cell_border

        # 多出来
        for row in sheet.iter_rows(min_row=12, values_only=False):
            for cell in row:
                sheet.row_dimensions[cell.row].height = row_height


        for ranges in ['B3:D3']:
            sheet.unmerge_cells(ranges)

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket['清关发票']


        new_sheet.cell(row=3, column=2, value=ticket_name[1])
        new_sheet.merge_cells(start_row=3, start_column=2, end_row=3,
                              end_column=4)


        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 模板G 百泰FBA17116ZL52-2023.02.16-德邦美森限时快船DG票-3件-美国-发票装箱单.xlsx
class debangmeisheng_usa(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['FBA专线出货资料模板']

        # 删除所有图片的方法
        sheet._images = []

        row_height = sheet.row_dimensions[12].height
        row_height_last = sheet.row_dimensions[15].height

        for r in range(9, 18):
            sheet.row_dimensions[r].height = sheet.row_dimensions[20].height

        # 这里要消除单元格
        for ranges in ['L9:L10', 'M9:M10', 'N9:N10', 'O9:O10',
                      'L11:L12', 'M11:M12', 'N11:N12', 'O11:O12',
                      'L13:L14', 'M13:M14', 'N13:N14', 'O13:O14','G4:H4']:
            sheet.unmerge_cells(ranges)

        # 删除原来的内容
        sheet.delete_rows(9, 18)

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket.active

        # 记录第17行的行高列宽，17行第一列的颜色和边框格式，和第15列的边框格式
        cell_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))
        cell_style = sheet.cell(row=12, column=2)
        cell_font = Font(name=cell_style.font.name, size=cell_style.font.size, bold=cell_style.font.bold,
                         italic=cell_style.font.italic,
                         vertAlign=cell_style.font.vertAlign, color=cell_style.font.color)
        cell_alignment = Alignment(horizontal='center', vertical='center')

        row_num = 9

        total_price = 0
        total_weigth = 0
        total_quantity =0
        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    new_sheet.cell(row=row_num, column=2, value=product.hs_code)
                    name = product.en_name + '(' + product.cn_name + ')'
                    new_sheet.cell(row=row_num, column=3, value=name)
                    new_sheet.cell(row=row_num, column=5, value=product.box_quantity[box_number])
                    new_sheet.cell(row=row_num, column=6, value=product.price)
                    new_sheet.cell(row=row_num, column=7, value=product.box_quantity[box_number] * int(product.price))
                    material = product.cn_material + '/' + product.en_material
                    new_sheet.cell(row=row_num, column=8, value=material)
                    usage = product.cn_usage + '/' + product.en_usage
                    new_sheet.cell(row=row_num, column=9, value=usage)
                    new_sheet.cell(row=row_num, column=10, value=product.has_magnetism)
                    new_sheet.cell(row=row_num, column=11, value=product.brand)
                    new_sheet.cell(row=row_num, column=12, value=box_number)

                    new_sheet.cell(row=row_num, column=13, value=box.weight)
                    new_sheet.cell(row=row_num, column=14, value=box.weight)
                    new_sheet.cell(row=row_num, column=15, value=box.box_message.length*box.box_message.width*box.box_message.height*0.000001)
                    new_sheet.cell(row=row_num, column=16, value=sku)
                    insert_images('D9', 'D', new_sheet, product_image, row_num, sku, row_height)
                    row_num += 1
                    total_weigth += box.weight
                    total_price +=  product.box_quantity[box_number]* product.price
                    total_quantity += product.box_quantity[box_number]
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=12, end_row=row_num - 1,
                                  end_column=12)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=13, end_row=row_num - 1,
                                  end_column=13)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=14, end_row=row_num - 1,
                                  end_column=14)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=15, end_row=row_num - 1,
                                  end_column=15)


        new_sheet.cell(row=8, column=16, value='SKU').border = cell_border
        new_sheet.cell(row=8, column=16, value='SKU').alignment = cell_alignment

        for row in new_sheet.iter_rows(min_row=9, min_col=2, max_col=16, max_row=row_num - 1, values_only=False):
            for cell in row:
                cell.border = cell_border
                cell.font = cell_font
                cell.alignment = cell_alignment
                new_sheet.row_dimensions[cell.row].height = row_height


        for row in new_sheet.iter_rows(min_row=row_num, min_col=2, max_col=16, max_row=row_num + 1, values_only=False):
            for cell in row:
                cell.border = cell_border

        new_sheet.row_dimensions[row_num].height = row_height_last
        new_sheet.row_dimensions[row_num + 1].height = row_height_last

        font_ticket = Font(name='Arial Unicode MS', size=12, bold=True)
        new_sheet.cell(row=4, column=7, value='FBA 号：' + ticket_name[1]).font = font_ticket
        new_sheet.merge_cells(start_row=4, start_column=7, end_row=4,
                              end_column=8)

        # total 行
        Arial_font = Font(name='Arial Unicode MS', size=12, bold=True)
        Arial_font_small = Font(name='Arial Unicode MS', size=12)
        new_sheet.cell(row=row_num + 1, column=2, value='TOTAL').font = Arial_font
        new_sheet.cell(row=row_num + 1, column=2, value='TOTAL').alignment = cell_alignment

        new_sheet.cell(row=row_num + 1, column=12, value=len(box_dict)).font = Arial_font_small
        new_sheet.cell(row=row_num + 1, column=12, value=len(box_dict)).alignment = cell_alignment

        new_sheet.cell(row=row_num + 1, column=13, value=total_weigth).font = Arial_font_small
        new_sheet.cell(row=row_num + 1, column=13, value=total_weigth).alignment = cell_alignment

        new_sheet.cell(row=row_num + 1, column=7, value=total_price).font = Arial_font_small
        new_sheet.cell(row=row_num + 1, column=7, value=total_price).alignment = cell_alignment

        new_sheet.cell(row=row_num + 1, column=5, value=total_quantity).font = Arial_font_small
        new_sheet.cell(row=row_num + 1, column=5, value=total_quantity).alignment = cell_alignment

        new_sheet.cell(row=row_num + 1, column=14, value=total_weigth).font = Arial_font_small
        new_sheet.cell(row=row_num + 1, column=14, value=total_weigth).alignment = cell_alignment

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 模板H-依诺达（新）深圳依诺达发票模板.xlsx
class shengzyiruoda_new(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['模板']

        row_height = sheet.row_dimensions[17].height


        cell_page = sheet.cell(row=17, column=17)
        cell_border = Border(left=cell_page.border.left, right=cell_page.border.right, top=cell_page.border.top,
                             bottom=cell_page.border.bottom)
        cell_style = sheet.cell(row=17, column=1)
        cell_font = Font(name=cell_style.font.name, size=cell_style.font.size, bold=cell_style.font.bold,
                         italic=cell_style.font.italic,
                         vertAlign=cell_style.font.vertAlign, color=cell_style.font.color)

        cell_Alignment = Alignment(horizontal='center', vertical='center')


        row_num = 17
        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    sheet.cell(row=row_num, column=1, value=box_number)
                    sheet.cell(row=row_num, column=2, value=box.weight)
                    sheet.cell(row=row_num, column=3, value=box.box_message.length)
                    sheet.cell(row=row_num, column=4, value=box.box_message.width)
                    sheet.cell(row=row_num, column=5, value=box.box_message.height)
                    sheet.cell(row=row_num, column=6, value=sku)
                    sheet.cell(row=row_num, column=7, value=product.en_name)
                    sheet.cell(row=row_num, column=8, value=product.cn_name)
                    sheet.cell(row=row_num, column=9, value=product.price)
                    sheet.cell(row=row_num, column=10, value=product.box_quantity[box_number])
                    sheet.cell(row=row_num, column=11, value=product.en_material)
                    sheet.cell(row=row_num, column=12, value=product.en_usage)
                    sheet.cell(row=row_num, column=13, value=product.hs_code)
                    sheet.cell(row=row_num, column=14, value=product.brand)
                    sheet.cell(row=row_num, column=15, value=product.model)
                    sheet.cell(row=row_num, column=16, value=product.link)
                    insert_images('Q17', 'Q', sheet, product_image, row_num, sku, row_height)

                    row_num += 1

        data_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        for row in sheet.iter_rows(min_row=16, min_col=1, max_col=22, values_only=False):
            for cell in row:
                cell.border = data_border
                cell.font = cell_font
                cell.alignment = cell_Alignment

        for row in sheet.iter_rows(min_row=17, min_col=17, max_col=17, values_only=False):
            for cell in row:
                cell.border = cell_border

        # 多出来
        for row in sheet.iter_rows(min_row=17, values_only=False):
            for cell in row:
                sheet.row_dimensions[cell.row].height = row_height

        for ranges in ['F1:H1', 'F2:H2','B15:D15','B1:D1']:
            sheet.unmerge_cells(ranges)

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket['模板']


        if '电' in ticket_name[2][0]:

            new_sheet.cell(row=1, column=6, value='是')
        else:
            new_sheet.cell(row=1, column=6, value='否')
        new_sheet.merge_cells(start_row=1, start_column=6, end_row=1,
                              end_column=8)

        if '磁' in ticket_name[2][0]:
            new_sheet.cell(row=2, column=6, value='是')
        else:
            new_sheet.cell(row=2, column=6, value='否')
        new_sheet.merge_cells(start_row=2, start_column=6, end_row=2,
                              end_column=8)


        new_sheet.cell(row=1, column=2, value=ticket_name[1])
        new_sheet.merge_cells(start_row=1, start_column=2, end_row=1,
                              end_column=4)

        new_sheet.cell(row=15, column=2, value=len(box_dict))
        new_sheet.merge_cells(start_row=15, start_column=2, end_row=15,
                              end_column=4)

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 模板I-百泰FBA171KZ8BL9-2023.03.03-林道UPS红单DB票-1件-美国-发票装箱单.xlsx，不用图片
class lindaoups_usa(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['发票']

        for ranges in ['A22:D22', 'A23:D23', 'A24:D24', 'A25:D25',
                       'A26:D26', 'A27:D27','C2:E2','B4:E4']:
            sheet.unmerge_cells(ranges)


        row_23_height = sheet.row_dimensions[23].height
        row_28_height = sheet.row_dimensions[28].height

        for r in range(19, 22):
            sheet.row_dimensions[r].height = sheet.row_dimensions[60].height

        # 记录E24单元格得的内容和格式
        cell_e24 = sheet.cell(row=24, column=5)
        cell_font_e24 = Font(name=cell_e24.font.name, size=cell_e24.font.size, bold=cell_e24.font.bold,
                             italic=cell_e24.font.italic,
                             vertAlign=cell_e24.font.vertAlign, color=cell_e24.font.color)

        cell_alignment_e24 = Alignment(horizontal=cell_e24.alignment.horizontal, vertical=cell_e24.alignment.vertical,
                                       text_rotation=cell_e24.alignment.text_rotation,
                                       wrap_text=cell_e24.alignment.wrap_text,
                                       shrink_to_fit=cell_e24.alignment.shrink_to_fit,
                                       indent=cell_e24.alignment.indent)

        sheet.delete_rows(22, 29)

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp_first.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket['发票']

        red_color = Color(rgb='FF0000')
        ticket_font = Font(name='Arial', size=11, bold=True, color=red_color)
        center_alignment = Alignment(horizontal='center', vertical='center')
        new_sheet.cell(row=2, column=3, value=ticket_name[1]).font = ticket_font
        new_sheet.cell(row=2, column=3).alignment = center_alignment
        new_sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=5)

        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        black_color = Color(rgb='000000')
        ticket_font_day = Font(name='Arial', size=11, bold=True, color=black_color)
        new_sheet.cell(row=4, column=2, value='DATE/日期'+formatted_date).font = ticket_font_day
        new_sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)


        # 记录第17行的行高列宽，17行第一列的颜色和边框格式，和第15列的边框格式
        row_height = sheet.row_dimensions[19].height
        cell_page = sheet.cell(row=19, column=1)
        cell_border = Border(left=cell_page.border.left, right=cell_page.border.right, top=cell_page.border.top,
                             bottom=cell_page.border.bottom)
        cell_style = sheet.cell(row=19, column=1)
        cell_font = Font(name=cell_style.font.name, size=cell_style.font.size, bold=cell_style.font.bold,
                         italic=cell_style.font.italic,
                         vertAlign=cell_style.font.vertAlign, color=cell_style.font.color)
        cell_Alignment = Alignment(horizontal='center', vertical='center')

        #这里的第一张需要记录总的，所以建立一个字典来装SKU 和sku对应数量和价格的总和
        total_sku = {}
        name_sku = []

        num_row = 19

        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0

                    print(total_sku.keys())
                    if product.en_name not in name_sku:
                        name = product.en_name + '(' + product.cn_name + ')'
                        total_sku[product.en_name] = [product.price,product.box_quantity[box_number],name]
                        name_sku.append(product.en_name)
                    else:
                        total_sku[product.en_name][0] += product.price
                        total_sku[product.en_name][1] += product.box_quantity[box_number]

        for index, (sku_index, value_sku) in enumerate(total_sku.items()):
                new_sheet.cell(row=num_row, column=1, value=value_sku[2]).font = cell_font
                new_sheet.cell(row=num_row, column=1, value=value_sku[2]).alignment = cell_Alignment
                new_sheet.cell(row=num_row, column=2, value=value_sku[1]).font = cell_font
                new_sheet.cell(row=num_row, column=2, value=value_sku[1]).alignment = cell_Alignment
                new_sheet.cell(row=num_row, column=3, value=value_sku[0]).font = cell_font
                new_sheet.cell(row=num_row, column=3, value=value_sku[0]).alignment = cell_Alignment
                new_sheet.cell(row=num_row, column=4, value=value_sku[0]*value_sku[1]).font = cell_font
                new_sheet.cell(row=num_row, column=4, value=value_sku[0]*value_sku[1]).alignment = cell_Alignment
                new_sheet.cell(row=num_row, column=5, value='CN').font = cell_font
                new_sheet.cell(row=num_row, column=5, value='CN').alignment = cell_Alignment
                num_row += 1

        # 设置行高和边框
        for row in range(19, int(num_row)):
            new_sheet.row_dimensions[row].height = row_height
            for col in range(1, 6):
                new_sheet.cell(row=row, column=col).border = cell_border

        for row in range(num_row, num_row + 8):
            for col in range(1, 6):
                new_sheet.cell(row=row, column=col).border = cell_border

        num = num_row
        new_sheet.cell(row=num, column=1,
                       value='THESE COMMODITIES ARE LICENSED FOR THE UNTIMATE DESTINATION SHOWN.').font = cell_font
        new_sheet.cell(row=num, column=1,
                       value='THESE COMMODITIES ARE LICENSED FOR THE UNTIMATE DESTINATION SHOWN.').alignment = cell_Alignment

        new_sheet.cell(row=num + 1, column=1, value='以上商品已有到最终目的地的许可。').font = cell_font
        new_sheet.cell(row=num + 1, column=1, value='以上商品已有到最终目的地的许可。').alignment = cell_Alignment
        new_sheet.row_dimensions[num + 1].height = row_23_height

        new_sheet.cell(row=num + 3, column=1,
                       value='I DECLARE ALL THE INFORMATION CONTAINED IN THIS INVOICE LIST TO BE TRUE AND CORRECT.').font = Font(
            name='Arial', size=9, color='000080')
        new_sheet.cell(row=num + 3, column=1,
                       value='I DECLARE ALL THE INFORMATION CONTAINED IN THIS INVOICE LIST TO BE TRUE AND CORRECT.').alignment = cell_Alignment

        new_sheet.cell(row=num + 4, column=1, value='以上申报均属实。').font = Font(name='宋体', size=11, color='FF0000',
                                                                                   bold=True)
        new_sheet.cell(row=num + 4, column=1, value='以上申报均属实。').alignment = cell_Alignment

        new_sheet.cell(row=num + 6, column=1,
                       value='SIGNATURE OF SHIPPER/EXPORTER(TYPE NAME TITLE AND SIGN):    ').font = Font(name='Arial',
                                                                                                         size=9,
                                                                                                         color='000080',
                                                                                                         bold=True)
        new_sheet.cell(row=num + 6, column=1,
                       value='SIGNATURE OF SHIPPER/EXPORTER(TYPE NAME TITLE AND SIGN):    ').alignment = cell_Alignment
        new_sheet.row_dimensions[num + 6].height = row_28_height

        new_sheet.cell(row=num + 7, column=1, value='寄件人/出口商签名( 正楷和职位) ').font = Font(name='宋体', size=9,
                                                                                                   color='000080',
                                                                                                   bold=True)
        new_sheet.cell(row=num + 7, column=1, value='寄件人/出口商签名( 正楷和职位) ').alignment = cell_Alignment

        new_sheet.cell(row=num, column=5, value='CHECK  ONE ').font = cell_font
        new_sheet.cell(row=num, column=5, value='CHECK  ONE').alignment = cell_Alignment

        new_sheet.cell(row=num + 1, column=5, value='□ F.O.B ').font = cell_font
        new_sheet.cell(row=num + 1, column=5, value='□ F.O.B').alignment = cell_Alignment

        new_sheet.cell(row=num + 2, column=5, value=cell_e24.value).font = cell_font_e24
        new_sheet.cell(row=num + 2, column=5).alignment = cell_alignment_e24

        new_sheet.cell(row=num + 6, column=4, value='DATE:').font = cell_font
        new_sheet.cell(row=num + 6, column=4, value='DATE:').alignment = cell_Alignment
        new_sheet.cell(row=num + 6, column=5, value=formatted_date).font = ticket_font_day

        new_sheet.cell(row=num + 7, column=4, value='日期').font = cell_font
        new_sheet.cell(row=num + 7, column=4, value='日期').alignment = cell_Alignment

        for row in range(num_row, num_row + 6):
            new_sheet.merge_cells(f'A{row}:D{row}')

        temp_path_mid = os.path.join(temp, 'temp_mid.xlsx')
        new_ticket.save(temp_path_mid)

        second_ticket = openpyxl.load_workbook(temp_path_mid)
        sheet_ticket = second_ticket['箱单']

        row_height_12 = sheet_ticket.row_dimensions[12].height
        cell_t = sheet_ticket.cell(row=12, column=3)
        cell_border = Border(left=cell_t.border.left, right=cell_t.border.right, top=cell_t.border.top,
                             bottom=cell_t.border.bottom)


        merged_cells = sheet_ticket.merged_cells
        cells_to_unmerge = []

        for merged_cell in merged_cells:
            if merged_cell.min_row >= 12:
                cells_to_unmerge.append(merged_cell)

        for cell_range in cells_to_unmerge:
            sheet_ticket.unmerge_cells(str(cell_range))

        # 删除旧值
        for row in range(12, 16):
            for col in range(1, 20):
                sheet_ticket.cell(row=row, column=col).value = None
        # sheet_ticket.delete_rows(12,14)

        temp_path_second = os.path.join(temp, 'temp_second.xlsx')
        second_ticket.save(temp_path_second)

        ticket_last = openpyxl.load_workbook(temp_path_second)
        second_sheet = ticket_last['箱单']


        total_length = 0
        total_weight = 0
        for obj in box_dict.values():
            total_length += len(obj.array)
            total_weight += obj.weight

        row_num_second = 12
        number_ticket = ticket_name[1] + 'U00000'
        #这里插入需要多少先不管
        second_sheet.insert_rows(12, total_length-3)

        word_font_14 = Font(name='宋体', size=14)
        word_font_11 = Font(name='Arial', size=11)


        #红色加粗的格式
        font_red = Font(color='FF0000', bold=True,name='Arial', size=14)

        #红色没加粗的格式
        font_red_light = Font(color='FF0000', bold=False, name='Arial', size=14)

        total_quantity = 0
        total_price = 0
        total_weight = 0

        product_atrribute = []
        if '电' in ticket_name[2][0]:
            product_atrribute.append('带电')
        else:
            product_atrribute.append('无电')
        if '磁' in ticket_name[2][0]:
            product_atrribute.append('带磁')
        else:
            product_atrribute.append('无磁')

        for i, (box_number, box) in enumerate(box_dict.items()):
            if i == 0:
                box_change = box_number
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    FBA_number = number_ticket + str(box_number)
                    second_sheet.cell(row=row_num_second, column=1, value=FBA_number).font = font_red
                    name = product.en_name + '/' + product.cn_name
                    second_sheet.cell(row=row_num_second, column=2, value=box_number).font = word_font_14
                    second_sheet.cell(row=row_num_second, column=3, value=product.cn_name).font = word_font_14
                    second_sheet.cell(row=row_num_second, column=4, value=product.en_name).font = word_font_14
                    second_sheet.cell(row=row_num_second, column=5, value=product.price).font = word_font_14
                    second_sheet.cell(row=row_num_second, column=6, value=product.box_quantity[box_number])
                    second_sheet.cell(row=row_num_second, column=7, value=product.en_material).font = font_red
                    second_sheet.cell(row=row_num_second, column=8, value=product.cn_usage).font = font_red
                    second_sheet.cell(row=row_num_second, column=9, value=product.model).font = font_red
                    second_sheet.cell(row=row_num_second, column=10, value=box.weight).font = font_red_light
                    second_sheet.cell(row=row_num_second, column=14, value=product.hs_code).font = font_red_light
                    insert_images('Q12', 'Q', second_sheet, product_image, row_num_second, sku, row_height_12)
                    second_sheet.cell(row=row_num_second, column=16, value=product_atrribute[0]+' '+product_atrribute[1]+' '+product.brand[:1]).font = font_red_light

                    total_weight += box.weight
                    total_price += product.price
                    total_quantity += product.box_quantity[box_number]

                    row_num_second += 1

                volume = str(box.box_message.length)+'*'+str(box.box_message.width)+'*'+str(box.box_message.height)

                second_sheet.cell(row=row_num_second - len(box.array), column=11, value= volume).font= font_red_light
            # second_sheet.cell(row=row_num_second - len(box.array), column=13, value=total_box_price)

            second_sheet.merge_cells(start_row=row_num_second - len(box.array), start_column=1, end_row=row_num_second - 1, end_column=1)
            second_sheet.merge_cells(start_row=row_num_second - len(box.array), start_column=10, end_row=row_num_second - 1, end_column=10)
            second_sheet.merge_cells(start_row=row_num_second - len(box.array), start_column=11, end_row=row_num_second - 1, end_column=13)
            second_sheet.merge_cells(start_row=row_num_second - len(box.array), start_column=2, end_row=row_num_second - 1,
                                  end_column=2)

        blue_font = Font(color='0000FF', bold=True,name='宋体', size=14)
        second_sheet.cell(row = row_num_second,column=1,value = '总件数' ).font = blue_font

        second_sheet.cell(row = row_num_second,column=2,value = len(box_dict)).font = blue_font

        second_sheet.cell(row = row_num_second,column=5,value = total_price).font = blue_font
        second_sheet.cell(row = row_num_second,column=6,value = total_quantity).font = blue_font
        second_sheet.cell(row = row_num_second,column=10,value = total_weight).font = blue_font


        row_height_second = second_sheet.row_dimensions[12].height
        for row in range(12, row_num_second):
            second_sheet.row_dimensions[row].height = row_height_second

        align = Alignment(horizontal='center', vertical='center')


        for rows in second_sheet.iter_rows(min_row=12, min_col=1, max_col=17, max_row=row_num_second-1, values_only=False):
            for cell in rows:
                cell.border = cell_border
                cell.alignment = align

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        ticket_last.save(result_path)
        os.remove(temp_path)
        os.remove(temp_path_second)
        os.remove(temp_path_mid)

# 模板J-百泰FBA171T3XY25-2023.03.07-UPS红单DJ票-2件-美国-发票装箱单.xlsx-不用图片
class hongdangDJ_usa(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['发票']

        for ranges in ['A13:A14', 'A15:A16', 'J13:J14', 'L13:L14', 'K13:K14', 'M13:M14', 'J15:J16', 'L15:L16',
                       'K15:K16', 'M15:M16',
                       'B13:B14', 'B15:B16', 'A19:F19', 'K19:O19', 'K20:O20','A7:C11','A4:C4']:
            sheet.unmerge_cells(ranges)

        row_height = sheet.row_dimensions[13].height
        cell_13 = sheet.cell(row=13, column=1)
        cell_font_13 = Font(name=cell_13.font.name, size=cell_13.font.size, bold=cell_13.font.bold,
                            italic=cell_13.font.italic,
                            vertAlign=cell_13.font.vertAlign, color=cell_13.font.color)

        # 删除原来的内容
        # sheet.delete_rows(13, 16)

        for r in range(13, 17):
            sheet.row_dimensions[r].height = sheet.row_dimensions[21].height

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket.active

        ticket_font = Font(name='宋体', size=12, bold=True)
        new_sheet.cell(row=4, column=1, value='运单号码：'+ticket_name[1]).font = ticket_font
        new_sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)

        new_sheet.cell(row=7, column=1, value= 'FBA：编号: ' + ticket_name[1]).font = ticket_font
        new_sheet.merge_cells(start_row=7, start_column=1, end_row=11, end_column=3)


        # 记录第17行的行高列宽，17行第一列的颜色和边框格式，和第15列的边框格式
        cell_page = sheet.cell(row=12, column=1)
        cell_border = Border(left=cell_page.border.left, right=cell_page.border.right, top=cell_page.border.top,
                             bottom=cell_page.border.bottom)
        cell_alignment = Alignment(horizontal='center', vertical='center')

        total_length = 0
        total_weight = 0

        for obj in box_dict.values():
            total_length += len(obj.array)
            total_weight += obj.weight

        row_num = 13
        number_ticket = ticket_name[1] + 'U00000'

        #这里有问题----干脆还是自己写吧
        new_sheet.insert_rows(13, total_length-4)

        for i, (box_number, box) in enumerate(box_dict.items()):
            total_quantity = 0
            total_price = 0
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    FBA_number = number_ticket + str(box_number)
                    new_sheet.cell(row=row_num, column=1, value=FBA_number).font = cell_font_13
                    new_sheet.cell(row=row_num, column=1, value=FBA_number).alignment = cell_alignment
                    # 这里有一个Reference ID必填？？？
                    # 这里由于限制先不填sku
                    volume = str(box.box_message.length) + '*' + str(box.box_message.width) + '*' + str(
                        box.box_message.height)
                    new_sheet.cell(row=row_num, column=2, value=box_number).font = cell_font_13
                    new_sheet.cell(row=row_num, column=2, value=box_number).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=3, value=product.cn_name).font = cell_font_13
                    new_sheet.cell(row=row_num, column=3, value=product.cn_name).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=4, value=product.en_name).font = cell_font_13
                    new_sheet.cell(row=row_num, column=4, value=product.en_name).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=5, value=product.price).font = cell_font_13
                    new_sheet.cell(row=row_num, column=5, value=product.price).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=6, value=product.box_quantity[box_number]).font = cell_font_13
                    new_sheet.cell(row=row_num, column=6, value=product.box_quantity[box_number]).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=7, value=product.box_quantity[box_number] * product.price).font = cell_font_13
                    new_sheet.cell(row=row_num, column=7,
                                   value=product.box_quantity[box_number] * product.price).alignment = cell_alignment

                    material = product.cn_material + '/' + product.en_material
                    new_sheet.cell(row=row_num, column=8, value=material).font = cell_font_13
                    new_sheet.cell(row=row_num, column=8, value=material).alignment = cell_alignment

                    usage = product.cn_usage + '/' + product.en_usage
                    new_sheet.cell(row=row_num, column=9, value=usage).font = cell_font_13
                    new_sheet.cell(row=row_num, column=9, value=usage).alignment = cell_alignment

                    # 这里是毛重----后面要商量一下
                    new_sheet.cell(row=row_num, column=10, value=box.weight).font = cell_font_13
                    new_sheet.cell(row=row_num, column=10, value=box.weight).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=11, value=box.box_message.length).font = cell_font_13
                    new_sheet.cell(row=row_num, column=11, value=box.box_message.length).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=12, value=box.box_message.width).font = cell_font_13
                    new_sheet.cell(row=row_num, column=12, value=box.box_message.width).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=13, value=box.box_message.height).font = cell_font_13
                    new_sheet.cell(row=row_num, column=13, value=box.box_message.height).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=14, value=product.brand).font = cell_font_13
                    new_sheet.cell(row=row_num, column=14, value=product.brand).alignment = cell_alignment

                    new_sheet.cell(row=row_num, column=15, value=product.hs_code).font = cell_font_13
                    new_sheet.cell(row=row_num, column=15, value=product.hs_code).alignment = cell_alignment
                    total_quantity += product.box_quantity[box_number]
                    total_price += product.quantity * product.price
                    row_num += 1
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=1, end_row=row_num - 1, end_column=1)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=2, end_row=row_num - 1, end_column=2)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=10, end_row=row_num - 1,
                                  end_column=10)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=11, end_row=row_num - 1,
                                  end_column=11)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=12, end_row=row_num - 1,
                                  end_column=12)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=13, end_row=row_num - 1,
                                  end_column=13)
            new_sheet.cell(row=row_num, column=2, value=len(box_dict))
            new_sheet.cell(row=row_num, column=6, value=total_quantity)
            new_sheet.cell(row=row_num, column=7, value=total_price)
            new_sheet.cell(row=row_num, column=10, value=total_weight)
            today = datetime.date.today()
            formatted_date = today.strftime("%Y.%m.%d")
            new_sheet.cell(row=row_num+3, column=11, value='签字日期:'+formatted_date)


        # 设置行高和边框
        for row in range(13, row_num):
            new_sheet.row_dimensions[row].height = row_height
            for col in range(1, 16):
                new_sheet.cell(row=row, column=col).border = cell_border


        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 模板k ，百泰FBA171F1JKW8-2023.03.01-德邦空派DG票-1件-美国-发票装箱单.xlsx
class gemDG_usa(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['箱单发票']

        cell_beagin = sheet.cell(row=9, column=3)
        cell_head = sheet.cell(row=7, column=2)
        cell_border = Border(left=cell_beagin.border.left, right=cell_beagin.border.right, top=cell_beagin.border.top,
                             bottom=cell_beagin.border.bottom)
        cell_font_9 = Font(name=cell_beagin.font.name, size=cell_beagin.font.size, bold=cell_beagin.font.bold,
                           italic=cell_beagin.font.italic,
                           vertAlign=cell_beagin.font.vertAlign, color=cell_beagin.font.color)
        cell_alignment = Alignment(horizontal='center', vertical='center')

        row_height = sheet.row_dimensions[9].height
        row_height_low = sheet.row_dimensions[16].height

        sheet.unmerge_cells('L4:M4')
        # 查找并取消合并单元格
        merged_cells = sheet.merged_cells
        cells_to_unmerge = []

        for merged_cell in merged_cells:
            if merged_cell.min_row >= 9:  # 从第9行开始
                cells_to_unmerge.append(merged_cell)

        for cell_range in cells_to_unmerge:
            sheet.unmerge_cells(str(cell_range))

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket.active

        total_length = 0
        total_weight = 0

        for obj in box_dict.values():
            total_length += len(obj.array)
            total_weight += obj.weight

        #插入时间
        current_date = datetime.datetime.now().strftime("%Y.%m.%d")
        new_sheet.cell(row=4,column=12,value=current_date)
        new_sheet.merge_cells(start_row=4, start_column=12, end_row=4,
                              end_column=13)

        # 插入表格格式的图片
        img = openpyxl.drawing.image.Image('相关信息/德邦快递图标.png')
        img.anchor = 'B3'
        new_sheet.add_image(img)


        row_num = 9
        number_ticket = ticket_name[1]+'U00000'
        new_sheet.insert_rows(9, total_length - 5)
        for i, (box_number, box) in enumerate(box_dict.items()):
            total_quantity = 0
            total_price = 0
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    FBA_number = number_ticket + str(box_number)
                    new_sheet.cell(row=row_num, column=2, value=FBA_number)
                    new_sheet.cell(row=row_num, column=3, value=box_number)
                    name = product.en_name + '(' + product.cn_name + ')'
                    new_sheet.cell(row=row_num, column=4, value=name)
                    new_sheet.cell(row=row_num, column=5, value=product.hs_code)
                    new_sheet.cell(row=row_num, column=6, value=product.box_quantity[box_number])
                    price = '$' + str(product.price)
                    new_sheet.cell(row=row_num, column=7, value=price)
                    t_price = '$' + str(round(product.box_quantity[box_number]* product.price, 2))
                    new_sheet.cell(row=row_num, column=8, value=t_price)
                    new_sheet.cell(row=row_num, column=9, value=box.weight)
                    new_sheet.cell(row=row_num, column=10, value=box.weight)
                    new_sheet.cell(row=row_num, column=11, value=box.weight)
                    new_sheet.cell(row=row_num, column=12, value=box.box_message.length)
                    new_sheet.cell(row=row_num, column=13, value=box.box_message.width)
                    new_sheet.cell(row=row_num, column=14, value=box.box_message.height)
                    new_sheet.cell(row=row_num, column=15, value=box.box_message.length * box.weight * box.box_message.height * 0.000001)
                    new_sheet.cell(row=row_num, column=17, value=product.has_magnetism)
                    insert_images('P9', 'P', new_sheet, product_image, row_num, sku, row_height)
                    total_quantity += product.box_quantity[box_number]
                    total_price += product.box_quantity[box_number] * product.price
                    row_num += 1

            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=1, end_row=row_num - 1, end_column=1)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=2, end_row=row_num - 1, end_column=2)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=12, end_row=row_num - 1,
                                  end_column=12)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=13, end_row=row_num - 1,
                                  end_column=13)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=14, end_row=row_num - 1,
                                  end_column=14)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=15, end_row=row_num - 1,
                                  end_column=15)
            new_sheet.cell(row=row_num, column=3, value=len(box_dict))
            new_sheet.cell(row=row_num, column=6, value=total_quantity)
            new_sheet.cell(row=row_num, column=8, value=total_price)
            new_sheet.cell(row=row_num, column=11, value=total_weight)

        left_border = cell_head.border.left
        thick_left = Border(left=left_border,
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))

        thick_last = Border(top=left_border,
                            bottom=left_border)

        thick_last_lefy = Border(left=left_border,
                                 top=left_border,
                                 bottom=left_border)
        thick_last_right = Border(
            right=left_border,
            top=left_border,
            bottom=left_border)
        thick_last_middle = Border(right=Side(border_style='thin'),
                                   left=Side(border_style='thin'),
                                   top=left_border,
                                   bottom=left_border)


        red_color = Color(rgb="FF0000")
        red_font = Font(color=red_color, name='微软雅黑', size=11)
        for row in new_sheet.iter_rows(min_row=9, min_col=2, max_col=18, max_row=row_num, values_only=False):
            for cell in row:
                if cell.column == 2:
                    if cell.row == row_num:
                        cell.border = thick_last_lefy
                        cell.font = cell_font_9
                        cell.alignment = cell_alignment
                    else:
                        cell.border = thick_left
                        cell.font = red_font
                        cell.alignment = cell_alignment
                elif cell.row == row_num:
                    if cell.column == 18:
                        cell.border = thick_last_right
                        cell.font = cell_font_9
                        cell.alignment = cell_alignment
                    elif cell.column == 15:
                        cell.border = thick_last_middle
                        cell.font = cell_font_9
                        cell.alignment = cell_alignment
                    else:
                        cell.border = thick_last
                        cell.font = cell_font_9
                        cell.alignment = cell_alignment
                else:
                    cell.border = cell_border
                    cell.font = cell_font_9
                    cell.alignment = cell_alignment

        for row in range(9, row_num):
            new_sheet.row_dimensions[row].height = row_height

        for i in range(row_num, row_num + 25):
            new_sheet.row_dimensions[i].height = row_height_low


        # 备注的合并
        new_sheet.merge_cells(start_row=row_num + 2, start_column=6, end_row=row_num + 8, end_column=13)

        #申报要素的填充
        new_sheet.merge_cells(start_row=row_num + 2, start_column=2, end_row=row_num + 2, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 10, start_column=2, end_row=row_num + 10, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 18, start_column=2, end_row=row_num + 18, end_column=4)

        #这里的申报要素是有几个产品就需要几个
        # 通过sku  确定有几个，然后根据行数自动生成


        new_sheet.merge_cells(start_row=row_num + 3, start_column=3, end_row=row_num + 3, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 4, start_column=3, end_row=row_num + 4, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 5, start_column=3, end_row=row_num + 5, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 6, start_column=3, end_row=row_num + 6, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 7, start_column=3, end_row=row_num + 7, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 8, start_column=3, end_row=row_num + 8, end_column=4)

        new_sheet.merge_cells(start_row=row_num + 11, start_column=3, end_row=row_num + 11, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 12, start_column=3, end_row=row_num + 12, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 13, start_column=3, end_row=row_num + 13, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 14, start_column=3, end_row=row_num + 14, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 15, start_column=3, end_row=row_num + 15, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 16, start_column=3, end_row=row_num + 16, end_column=4)

        new_sheet.merge_cells(start_row=row_num + 19, start_column=3, end_row=row_num + 19, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 20, start_column=3, end_row=row_num + 20, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 21, start_column=3, end_row=row_num + 21, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 22, start_column=3, end_row=row_num + 22, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 23, start_column=3, end_row=row_num + 23, end_column=4)
        new_sheet.merge_cells(start_row=row_num + 24, start_column=3, end_row=row_num + 24, end_column=4)

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 模板L 百泰FBA15GMRY8YG-2023.01.31-UPS红单KU票-1件-英国-发票装箱单.xlsx
class hongdangKU_eng(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['EU 随货发票参考']

        row_height = sheet.row_dimensions[17].height

        cell_t = sheet.cell(row=16, column=6)

        cell_border = Border(left=cell_t.border.left, right=cell_t.border.right, top=cell_t.border.top,
                             bottom=cell_t.border.bottom)

        # 查找并取消合并单元格
        merged_cells = sheet.merged_cells
        cells_to_unmerge = []

        for merged_cell in merged_cells:
            if merged_cell.min_row >= 17:
                cells_to_unmerge.append(merged_cell)

        for cell_range in cells_to_unmerge:
            sheet.unmerge_cells(str(cell_range))

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket.active

        word_font_12 = Font(name='宋体', size=12)
        word_font_11 = Font(name='宋体', size=11)

        total_length = 0
        total_weight = 0
        for obj in box_dict.values():
            total_length += len(obj.array)
            total_weight += obj.weight

        row_num = 17
        number_ticket = ticket_name[1]+'U00000'
        new_sheet.insert_rows(17, total_length - 3)
        for i, (box_number, box) in enumerate(box_dict.items()):
            total_quantity = 0
            total_price = 0
            total_box_price = 0
            if i == 0:
                box_change = box_number
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    FBA_number = number_ticket + str(box_number)
                    new_sheet.cell(row=row_num, column=1, value=FBA_number).font = word_font_12
                    name = product.en_name + '/' + product.cn_name
                    new_sheet.cell(row=row_num, column=2, value=name).font = word_font_11
                    new_sheet.cell(row=row_num, column=6, value=product.hs_code).font = word_font_11
                    new_sheet.cell(row=row_num, column=7, value=product.en_usage).font = word_font_11
                    new_sheet.cell(row=row_num, column=8, value=product.en_material).font = word_font_11
                    new_sheet.cell(row=row_num, column=9, value='China')
                    new_sheet.cell(row=row_num, column=10, value=product.box_quantity[box_number])
                    new_sheet.cell(row=row_num, column=11, value=product.price)
                    new_sheet.cell(row=row_num, column=12, value=product.price * product.box_quantity[box_number])
                    total_box_price += product.box_quantity[box_number] * product.price
                    total_quantity += product.box_quantity[box_number]
                    total_price += product.box_quantity[box_number] * product.price

                    new_sheet.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
                    row_num += 1

            new_sheet.cell(row=row_num - len(box.array), column=13, value=total_box_price)
            total_box_price = 0
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=1, end_row=row_num - 1, end_column=1)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=13, end_row=row_num - 1,
                                  end_column=14)

        # 赋值
        new_sheet.cell(row=row_num, column=12).value = total_price
        new_sheet.cell(row=row_num + 1, column=10).value = total_quantity

        # 合并单元格格式
        new_sheet.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=9)
        new_sheet.merge_cells(start_row=row_num, start_column=13, end_row=row_num, end_column=14)
        new_sheet.merge_cells(start_row=row_num + 1, start_column=13, end_row=row_num + 1, end_column=14)
        new_sheet.merge_cells(start_row=row_num + 1, start_column=2, end_row=row_num + 1, end_column=5)

        new_sheet.merge_cells(start_row=row_num + 2, start_column=13, end_row=row_num + 2, end_column=14)
        new_sheet.merge_cells(start_row=row_num + 2, start_column=13, end_row=row_num + 3, end_column=14)
        new_sheet.merge_cells(start_row=row_num + 2, start_column=2, end_row=row_num + 3, end_column=11)
        new_sheet.merge_cells(start_row=row_num + 2, start_column=12, end_row=row_num + 3, end_column=12)

        new_sheet.merge_cells(start_row=row_num + 5, start_column=2, end_row=row_num + 6, end_column=10)
        new_sheet.merge_cells(start_row=row_num + 9, start_column=1, end_row=row_num + 10, end_column=14)
        new_sheet.merge_cells(start_row=row_num + 12, start_column=1, end_row=row_num + 20, end_column=10)

        for row in range(17, row_num + 2):
            new_sheet.row_dimensions[row].height = row_height

        align = Alignment(horizontal='center', vertical='center')

        for row in new_sheet.iter_rows(min_row=17, min_col=1, max_col=14, max_row=row_num + 3, values_only=False):
            for cell in row:
                cell.border = cell_border
                cell.alignment = align

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)


'''下面三个是日本的模板，数据源会有不同'''
# 模板M 递信 JPE-JP-FBA15DCLNMJP海外仓贴标资料对应.xlsx
class dixing_jap(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):

        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb.active

        row_height = sheet.row_dimensions[3].height

        # 查找并取消合并单元格
        merged_cells = sheet.merged_cells
        cells_to_unmerge = []

        for merged_cell in merged_cells:
            if merged_cell.min_row >= 3:  # 从第9行开始
                cells_to_unmerge.append(merged_cell)

        for cell_range in cells_to_unmerge:
            sheet.unmerge_cells(str(cell_range))

        last_heigth = sheet.row_dimensions[21].height

        # 删除原来的内容
        sheet.delete_rows(3, 20)

        for r in range(3, 21):
            sheet.row_dimensions[r].height = sheet.row_dimensions[25].height

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket.active


        get_ticket_pattern = r'[a-zA-Z\s]+票'
        output_name = creat_output_name(ticket_name[2]).split('-')[1]
        match_obj = re.search(get_ticket_pattern, output_name)
        if match_obj:
            result = match_obj.group()
            result = result[:-1]  # 去掉 "票" 字符
        else:
            print("No match found")


        total_length = 0
        total_weight = 0
        total_quantity = 0
        total_number = 0
        for obj in box_dict.values():
            total_length += len(obj.array)
            if obj.weight is not None:
                total_weight += obj.weight

        row_num = 3
        ticket = ticket_name[1]+'U00000'

        # 格式，边框，和最后一行

        center_alignment = Alignment(horizontal='center', vertical='center')

        for i, (box_number, box) in enumerate(box_dict.items()):

            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):

                    if sku not in product_dict.keys():
                        print("产品导出中没有这个sku"+sku)
                        continue
                    else:
                        product = product_dict[sku]
                        if product.price == None or product.price == '':
                            product.price = 0

                    identifier = result +'0'+ str(box_number)
                    new_sheet.cell(row=row_num, column=1, value=identifier).alignment = center_alignment
                    name = product.cn_name + '\n' + product.en_name
                    new_sheet.cell(row=row_num, column=2, value=name).alignment = center_alignment
                    material = product.cn_material + '\n' + product.en_material
                    new_sheet.cell(row=row_num, column=3, value=material).alignment = center_alignment
                    usage = product.en_usage + '，' + product.cn_usage
                    new_sheet.cell(row=row_num, column=4, value=usage).alignment = center_alignment
                    new_sheet.cell(row=row_num, column=5, value=1).alignment = center_alignment
                    new_sheet.cell(row=row_num, column=6, value=box.weight).alignment = center_alignment

                    insert_images('L3', 'L', new_sheet, product_image, row_num, sku, row_height)
                    chichun = str(box.box_message.length) + '*' + str(box.box_message.width) + '*' + str(
                        box.box_message.height)
                    new_sheet.cell(row=row_num, column=7, value=chichun).alignment = center_alignment

                    current_date = datetime.datetime.now().strftime("%Y.%m.%d")
                    new_sheet.cell(row=1, column=4, value=current_date)

                    if box_number in product.box_quantity.keys():
                        real_quantity = str(product.box_quantity[box_number]).split(' ')
                        # if len(real_quantity) < 2:
                        #     raise ValueError(f"Unsupported : {real_quantity}")
                    # 这里有疑问，这里指的是一个箱子中的总数量还是一个这个产品再箱子里面的重量


                    new_sheet.cell(row=row_num, column=9, value=real_quantity[1]).alignment = center_alignment
                    new_sheet.cell(row=row_num, column=13, value= str(ticket_name[1])+'U'+'00000'+str(box_number)).alignment = center_alignment


                    match = re.match(r'^([A-Za-z])(\d)$', real_quantity[0])

                    new_sheet.cell(row=row_num, column=10,
                                   value="{}{}".format(match.group(1), match.group(2))+' '+ real_quantity[1]).alignment = center_alignment

                    total_quantity += int(real_quantity[1])
                    total_number += int(real_quantity[1])
                    row_num += 1

            new_sheet.cell(row=row_num- len(box.array), column=8, value=total_number).alignment = center_alignment
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=1, end_row=row_num - 1, end_column=1)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=5, end_row=row_num - 1, end_column=5)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=6, end_row=row_num - 1, end_column=6)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=7, end_row=row_num - 1, end_column=7)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=8, end_row=row_num - 1, end_column=8)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=13, end_row=row_num - 1,end_column=13)
            total_number = 0
        new_sheet.merge_cells(start_row=3, start_column=14, end_row=row_num, end_column=14)



        new_sheet.delete_rows(row_num, row_num + 18)

        cell_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        # 设置行高
        for row in range(3, row_num):
            new_sheet.row_dimensions[row].height = row_height

        for row in new_sheet.iter_rows(min_row=3, min_col=1, max_col=14, max_row=row_num, values_only=False):
            for cell in row:
                cell.border = cell_border

        new_sheet.row_dimensions[row_num].height = last_heigth
        last_font = Font(bold=True, name='宋体', size=12)

        new_sheet.cell(row=row_num, column=1, value='汇总').font = last_font
        new_sheet.cell(row=row_num, column=1, value='汇总').alignment = center_alignment

        last_font_five = Font(bold=True, name='微软雅黑', size=11)

        new_sheet.cell(row=row_num, column=5, value=len(box_dict)).font = last_font_five
        new_sheet.cell(row=row_num, column=5, value=len(box_dict)).alignment = center_alignment

        last_font_six = Font(bold=True, name='微软雅黑', size=9)
        new_sheet.cell(row=row_num, column=6, value=total_weight).font = last_font_six
        new_sheet.cell(row=row_num, column=6, value=total_weight).alignment = center_alignment

        last_font_eight = Font(bold=True, name='微软雅黑', size=12)
        new_sheet.cell(row=row_num, column=8, value=total_quantity).font = last_font_eight
        new_sheet.cell(row=row_num, column=8, value=total_quantity).alignment = center_alignment

        last_font_night = Font(bold=True, name='微软雅黑', size=11)
        new_sheet.cell(row=row_num, column=9, value=total_quantity).font = last_font_night
        new_sheet.cell(row=row_num, column=9, value=total_quantity).alignment = center_alignment

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)

# 模板N 宏川
class hongchuang_jap(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['装箱明细单']

        def read_data(file_name, sheet_name, column_names):
            data_dict = {}
            column_indices = {}
            wb = openpyxl.load_workbook(file_name)
            sheet = wb[sheet_name]
            header_row = sheet[1]
            for i, cell in enumerate(header_row):
                column_name = cell.value
                if column_name in column_names:
                    column_indices[column_name] = i + 1
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                box_num = row[column_indices['NO.'] - 1]
                product_id = row[column_indices['产品标题'] - 1]
                box_identity_obj = box_identity(
                    box_num=box_num,
                    inner_self=row[column_indices['自编内标'] - 1],
                    inner_FN=row[column_indices['亚马逊FNSKU'] - 1],
                    product_id=product_id,
                    box_quantity=row[column_indices['数量'] - 1],
                    inner_product=row[column_indices['自编货件号'] - 1]
                )
                if product_id not in data_dict:
                    data_dict[product_id] = box_identity_obj
                else:
                    continue
            return data_dict

        # 使用示例：
        column_names = ['NO.', '自编内标', '亚马逊FNSKU', '产品标题', '数量', '自编货件号']
        path_file = os.path.join(message_folder,model_file_name)
        data_dict = read_data(path_file, '换标表格', column_names)

        no_border = Border(left=Side(style='none'),
                           right=Side(style='none'),
                           top=Side(style='none'),
                           bottom=Side(style='none'))


        # 删除旧值
        for row in range(9, 25):
            for col in range(1, 5):
                sheet.cell(row=row, column=col).value = None
                sheet.cell(row=row, column=col).border = no_border

        #  下面是第一张
        total_length = sum(len(obj.array) for obj in box_dict.values())
        total_length += len(box_dict) - 1  # add number of blank rows

        sheet.cell(row = 3, column  = 2,value =ticket_name[1])
        row_num = 9
        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    identifier = 'Ja' + '-' + str(box_number)
                    sheet.cell(row=row_num, column=1, value=identifier)
                    sheet.cell(row=row_num, column=2, value=data_dict[sku].inner_self)
                    sheet.cell(row=row_num, column=3, value=data_dict[sku].inner_FN)
                    sheet.cell(row=row_num, column=4, value=data_dict[sku].box_num)
                    row_num += 1

                if row_num != total_length + 8:  # add blank row if not the last row
                    sheet.cell(row=row_num, column=1, value='')
                    sheet.cell(row=row_num, column=2, value='')
                    sheet.cell(row=row_num, column=3, value='')
                    sheet.cell(row=row_num, column=4, value='')
                    row_num += 1

        # 边框,字体，居中
        cell_t = sheet.cell(row=8, column=1)
        cell_border = Border(left=cell_t.border.left, right=cell_t.border.right, top=cell_t.border.top,
                             bottom=cell_t.border.bottom)
        align = Alignment(horizontal='center', vertical='center')
        word_font = Font(name='等线', size=11)

        for row in sheet.iter_rows(min_row=9, min_col=1, max_col=4, max_row=row_num - 2, values_only=False):

            for cell in row:
                cell.border = cell_border
                cell.alignment = align
                cell.font = word_font

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp_first.xlsx')
        wb.save(temp_path)

        second_ticket = openpyxl.load_workbook(temp_path)
        sheet_ticket = second_ticket['发票']

        # 下面是第二张
        # 取消合并单元格
        merged_cells = sheet_ticket.merged_cells
        cells_to_unmerge = []

        for merged_cell in merged_cells:
            if merged_cell.min_row >= 6:  # 从第9行开始
                cells_to_unmerge.append(merged_cell)

        for cell_range in cells_to_unmerge:
            sheet_ticket.unmerge_cells(str(cell_range))

        # 删除旧值
        for row in range(6, 16):
            for col in range(3, 21):
                sheet_ticket.cell(row=row, column=col).value = None


        temp_path_second = os.path.join(temp, 'temp_second.xlsx')
        second_ticket.save(temp_path_second)

        new_ticket = openpyxl.load_workbook(temp_path_second)
        new_sheet = new_ticket['发票']


        row_height = new_sheet.row_dimensions[6].height

        row_height_last = new_sheet.row_dimensions[16].height
        row_height_last_small = new_sheet.row_dimensions[17].height

        # 删除所有图片的方法，好像目前只有这个有效
        new_sheet._images = []

        total_length_ticket = 0
        total_weight = 0

        for obj in box_dict.values():
            total_length_ticket += len(obj.array)
            if obj.weight is not None:
                total_weight += obj.weight

        row_num_new = 6

        # 格式，边框，和最后一行

        center_alignment = Alignment(horizontal='center', vertical='center')
        new_sheet.insert_rows(6, total_length_ticket - 9)

        new_sheet.cell(row=5, column=9, value='中英文材质').alignment = center_alignment
        new_sheet.cell(row=5, column=9, value='中英文材质').font = Font(name='等线', size=12)

        total_quantity = 0

        # 定义一个字典，用于记录每个box_number对应的第一个sku
        sku_dict = {}
        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    # 如果当前sku对应的box_number已经在sku_dict中存在，则不需要再赋值identifier
                    if sku in sku_dict and sku_dict[sku] == box_number:
                        identifier = None
                    else:
                        # 否则，将当前sku对应的box_number加入sku_dict，并赋值identifier
                        sku_dict[sku] = box_number
                        if sku in data_dict:
                            identifier = data_dict[sku].inner_product + ' ' + str(box_number) + '箱'
                            new_sheet.cell(row=row_num_new, column=3, value=identifier).alignment = center_alignment
                        else:
                            print('sku没找到')
                    # 其他列的赋值保持不变
                    new_sheet.cell(row=row_num_new, column=4, value=product.en_name).alignment = center_alignment
                    new_sheet.cell(row=row_num_new, column=5, value=product.cn_name).alignment = center_alignment
                    new_sheet.cell(row=row_num_new, column=7, value=product.price).alignment = center_alignment
                    new_sheet.cell(row=row_num_new, column=6,
                                   value=product.box_quantity[box_number]).alignment = center_alignment
                    new_sheet.cell(row=row_num_new, column=8,
                                   value=product.box_quantity[box_number] * product.price).alignment = center_alignment
                    material = product.en_material + '\n' + product.cn_material
                    new_sheet.cell(row=row_num_new, column=9, value=material).alignment = center_alignment
                    new_sheet.cell(row=row_num_new, column=10, value=product.cn_usage).alignment = center_alignment

                    insert_images('K6', 'K', new_sheet, product_image, row_num, sku, row_height)
                    total_quantity += product.box_quantity[box_number]
                    row_num_new += 1

            # new_sheet.merge_cells(start_row=row_num_new - len(box.array), start_column=3, end_row=row_num - 1, end_column=3)

        # 边框
        cell_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        for row in range(6, row_num_new):
            new_sheet.row_dimensions[row].height = row_height

        # 中间有空着一行要注意
        new_sheet.row_dimensions[row_num_new].height = row_height_last

        for row in range(row_num_new + 1, row_num_new + 5):
            new_sheet.row_dimensions[row].height = row_height_last_small

        new_sheet.delete_rows(row_num_new)

        new_sheet.cell(row=row_num_new + 2, column=6, value=total_quantity).alignment = center_alignment

        for row in new_sheet.iter_rows(min_row=5, min_col=3, max_col=11, max_row=row_num_new, values_only=False):
            for cell in row:
                cell.border = cell_border


        temp_path_third = os.path.join(temp, 'temp_third.xlsx')
        new_ticket.save(temp_path_third)

        third_ticket = openpyxl.load_workbook(temp_path_third)
        sheet_third = third_ticket['海运装箱单']

        row_height = sheet_third.row_dimensions[6].height
        row_height_16 = sheet_third.row_dimensions[16].height
        row_height_normal = sheet_third.row_dimensions[40].height

        # 取消合并单元格

        merged_cells = sheet_third.merged_cells
        cells_to_unmerge = []

        for merged_cell in merged_cells:
            if merged_cell.min_row >= 6:  # 从第6行开始
                cells_to_unmerge.append(merged_cell)

        for cell_range in cells_to_unmerge:
            sheet_third.unmerge_cells(str(cell_range))

        # 移除多余的格式，不让哪一行的格式是固定的
        for col in range(1, sheet_third.max_column + 1):
            cell = sheet_third.cell(row=25, column=col)

            # 移除背景颜色
            cell.fill = PatternFill(fill_type=None)

            # 移除边框
            no_border = Side(style=None)
            cell.border = Border(left=no_border, right=no_border, top=no_border, bottom=no_border)

        for row in range(6, 16):
            sheet_third.row_dimensions[row].height = row_height_normal

        # 删除所有图片的方法，好像目前只有这个有效
        sheet_third._images = []


        temp_path_last= os.path.join(temp, 'last.xlsx')
        third_ticket.save(temp_path_last)

        new_ticket_third = openpyxl.load_workbook(temp_path_last)
        new_sheet_last= new_ticket_third['海运装箱单']


        # 删除旧值
        for row in range(6, 37):
            for col in range(1, 16):
                new_sheet_last.cell(row=row, column=col).value = None

        total_length_ticket_last = 0
        total_weight = 0
        for obj in box_dict.values():
            total_length_ticket_last += len(obj.array)
            if obj.weight is not None:
                total_weight += obj.weight

        row_num_new = 6

        # 字体先欠着
        Arial_font = Font(name='Arial', size=12)
        deng_font = Font(name='等线', size=12)
        song_font = Font(name='华文仿宋', size=12)
        song_font_bold = Font(name='华文仿宋', size=12, bold=True)

        # 定义一个字典，用于记录每个box_number对应的第一个sku
        sku_dict_last = {}

        total_quantity = 0
        total_weight = 0
        for i, (box_number, box) in enumerate(box_dict.items()):
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    # 如果当前sku对应的box_number已经在sku_dict中存在，则不需要再赋值identifier
                    if sku in sku_dict_last and sku_dict_last[sku] == box_number:
                        identifier = None
                    else:
                        # 否则，将当前sku对应的box_number加入sku_dict，并赋值identifier
                        sku_dict_last[sku] = box_number
                        if sku in data_dict:
                            identifier = data_dict[sku].inner_product + str(len(box_dict)) + '箱' + ' (' + str(
                                box_number) + ' )'
                            new_sheet_last.cell(row=row_num_new, column=1,
                                                value=identifier).alignment = center_alignment
                            new_sheet_last.cell(row=row_num_new, column=1, value=identifier).font = Arial_font

                        else:
                            print('sku没找到')
                    # 其他列的赋值保持不变
                    new_sheet_last.cell(row=row_num_new, column=2, value=product.en_name).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=2, value=product.en_name).font = Arial_font

                    new_sheet_last.cell(row=row_num_new, column=3, value=product.cn_name).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=3, value=product.cn_name).font = deng_font

                    new_sheet_last.cell(row=row_num_new, column=4, value=1).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=4, value=1).font = deng_font

                    new_sheet_last.cell(row=row_num_new, column=5,
                                        value=product.box_quantity[box_number]).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=5,
                                        value=product.box_quantity[box_number]).font = deng_font

                    material = product.en_material + '\n' + product.cn_material
                    new_sheet_last.cell(row=row_num_new, column=6, value=material).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=6, value=material).font = deng_font

                    # 这里的重要要检查一下
                    new_sheet_last.cell(row=row_num_new, column=7, value=box.weight).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=7, value=box.weight).font = Arial_font

                    new_sheet_last.cell(row=row_num_new, column=8, value=box.weight).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=8, value=box.weight).font = Arial_font

                    new_sheet_last.cell(row=row_num_new, column=9, value=box.box_message.length*box.weight*box.box_message.height*0.000001).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=9, value=box.box_message.length*box.weight*box.box_message.height*0.000001).font = song_font

                    new_sheet_last.cell(row=row_num_new, column=10, value=product.cn_usage).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=10, value=product.cn_usage).font = song_font

                    new_sheet_last.cell(row=row_num_new, column=11,
                                        value=product.has_magnetism).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=11, value=product.has_magnetism).font = song_font_bold

                    new_sheet_last.cell(row=row_num_new, column=12, value=product.hs_code).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=12, value=product.hs_code).font = Arial_font

                    new_sheet_last.cell(row=row_num_new, column=14, value=product.model).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=14, value=product.model).font = song_font_bold

                    new_sheet_last.cell(row=row_num_new, column=15, value=product.brand).alignment = center_alignment
                    new_sheet_last.cell(row=row_num_new, column=15, value=product.brand).font = song_font_bold


                    insert_images('M6', 'M', new_sheet_last, product_image, row_num, sku, row_height)

                    total_quantity += product.box_quantity[box_number]
                    total_weight += box.weight
                    row_num_new += 1

            new_sheet_last.merge_cells(start_row=row_num_new - len(box.array), start_column=1, end_row=row_num_new - 1,
                                       end_column=1)
            new_sheet_last.merge_cells(start_row=row_num_new - len(box.array), start_column=4, end_row=row_num_new - 1,
                                       end_column=4)
            new_sheet_last.merge_cells(start_row=row_num_new - len(box.array), start_column=7, end_row=row_num_new - 1,
                                       end_column=7)
            new_sheet_last.merge_cells(start_row=row_num_new - len(box.array), start_column=8, end_row=row_num_new - 1,
                                       end_column=8)
            new_sheet_last.merge_cells(start_row=row_num_new - len(box.array), start_column=9, end_row=row_num_new - 1,
                                       end_column=9)


        cell_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))

        for row in range(6, row_num_new):
            new_sheet_last.row_dimensions[row].height = row_height

        new_sheet_last.row_dimensions[row_num_new].height = row_height_16
        for row in new_sheet_last.iter_rows(min_row=6, min_col=1, max_col=16, max_row=row_num_new, values_only=False):
            for cell in row:
                cell.border = cell_border

        new_sheet_last.cell(row=row_num_new, column=1, value='Total').font = deng_font
        new_sheet_last.cell(row=row_num_new, column=4, value=len(box_dict)).font = deng_font
        new_sheet_last.cell(row=row_num_new, column=5, value=total_quantity).font = deng_font
        new_sheet_last.cell(row=row_num_new, column=7, value=total_weight).font = deng_font
        new_sheet_last.cell(row=row_num_new, column=8, value=total_weight).font = deng_font

        new_sheet_last.cell(row=row_num_new, column=1, value='Total').alignment = center_alignment
        new_sheet_last.cell(row=row_num_new, column=4, value=len(box_dict)).alignment = center_alignment
        new_sheet_last.cell(row=row_num_new, column=5, value=total_quantity).alignment = center_alignment
        new_sheet_last.cell(row=row_num_new, column=7, value=total_weight).alignment = center_alignment
        new_sheet_last.cell(row=row_num_new, column=8, value=total_weight).alignment = center_alignment

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket_third.save(result_path)
        os.remove(temp_path)
        os.remove(temp_path_second)
        os.remove(temp_path_third)
        os.remove(temp_path_last)

#林道日本
class lindao_jap(InvoiceTemplate):
    def fill_data(self, dict_box, dict_product, model_file_name,ticket_name):
        wb = openpyxl.load_workbook(model_file_name)
        sheet = wb['日本发票']

        row_height = sheet.row_dimensions[8].height

        # for r in range(8, 14):
        #     sheet.row_dimensions[r].height = sheet.row_dimensions[21].height

        cell_t = sheet.cell(row=8, column=1)

        cell_border = Border(left=cell_t.border.left, right=cell_t.border.right, top=cell_t.border.top,
                             bottom=cell_t.border.bottom)

        #表头信息处理
        sheet.unmerge_cells(start_row=2, start_column=17, end_row=3, end_column=18)
        sheet.unmerge_cells(start_row=4, start_column=17, end_row=5, end_column=18)

        # 查找并取消合并单元格
        merged_cells = sheet.merged_cells
        cells_to_unmerge = []

        for merged_cell in merged_cells:
            if  merged_cell.min_row >= 8:
                cells_to_unmerge.append(merged_cell)

        for cell_range in cells_to_unmerge:
            sheet.unmerge_cells(str(cell_range))

        temp = os.path.join(os.getcwd(), '相关信息')
        temp_path = os.path.join(temp, 'temp.xlsx')
        wb.save(temp_path)

        new_ticket = openpyxl.load_workbook(temp_path)
        new_sheet = new_ticket.active

        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        new_sheet.cell(row=2, column=17, value=formatted_date)
        new_sheet.merge_cells(start_row=2, start_column=17, end_row=3,
                              end_column=18)

        new_sheet.cell(row=4, column=17, value = ticket_name[1]+'-')
        new_sheet.merge_cells(start_row=4, start_column=17, end_row=5,
                              end_column=18)

        word_font_Arial = Font(name='Arial', size=10)

        total_length = 0
        total_weight = 0
        total_vol = 0
        for obj in box_dict.values():
            total_length += len(obj.array)
            total_weight += obj.weight

        total_quantity = 0
        total_weight = 0
        total_box_price = 0

        row_num = 8
        number_ticket = ticket_name[1]+'U00000'
        new_sheet.insert_rows(8, total_length-1)
        new_ticket.save(temp_path)
        for i, (box_number, box) in enumerate(box_dict.items()):

            if i == 0:
                box_change = box_number
            if len(box.array) == 0:
                continue
            else:
                for j, sku in enumerate(box.array):
                    product = product_dict[sku]
                    if product.price == None or product.price == '':
                        product.price = 0
                    FBA_number = number_ticket + str(box_number)
                    new_sheet.cell(row=row_num, column=1, value=FBA_number).font = word_font_Arial
                    name = product.en_name + product.cn_name
                    new_sheet.cell(row=row_num, column=2, value=sku).font = word_font_Arial
                    insert_images('C8', 'C', new_sheet, product_image, row_num, sku, row_height)
                    new_sheet.cell(row=row_num, column=4, value=product.link).font = word_font_Arial
                    new_sheet.cell(row=row_num, column=5, value=name).font = word_font_Arial
                    new_sheet.cell(row=row_num, column=6, value=product.cn_material+'\n'+product.en_material).font = word_font_Arial
                    new_sheet.cell(row=row_num, column=7, value=product.cn_usage + '' + product.en_usage).font = word_font_Arial
                    new_sheet.cell(row=row_num, column=8, value=product.box_quantity[box_number]).font = word_font_Arial

                    new_sheet.cell(row=row_num, column=18, value=product.hs_code).font = word_font_Arial
                    new_sheet.cell(row=row_num, column=17, value=product.brand).font = word_font_Arial

                    total_quantity += product.box_quantity[box_number]

                    row_num += 1

            new_sheet.cell(row=row_num - len(box.array), column=9, value=box.weight)
            new_sheet.cell(row=row_num - len(box.array), column=10, value=1)   # 这里的总箱数要问一下
            new_sheet.cell(row=row_num - len(box.array), column=11, value=box.box_message.length)
            new_sheet.cell(row=row_num - len(box.array), column=12, value=box.box_message.width)
            new_sheet.cell(row=row_num - len(box.array), column=13, value=box.box_message.height)
            new_sheet.cell(row=row_num - len(box.array), column=14, value=box.box_message.length*box.box_message.width*box.box_message.height*0.000001)

            total_weight += box.weight
            total_vol += box.box_message.length*box.box_message.width*box.box_message.height*0.000001

            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=9, end_row=row_num - 1,end_column=9)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=10, end_row=row_num - 1,
                                  end_column=10)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=11, end_row=row_num - 1,
                                  end_column=11)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=12, end_row=row_num - 1,
                                  end_column=12)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=13, end_row=row_num - 1,
                                  end_column=13)
            new_sheet.merge_cells(start_row=row_num - len(box.array), start_column=14, end_row=row_num - 1,
                                  end_column=14)

        new_sheet.cell(row=row_num,column =10, value= len(box_dict))
        new_sheet.cell(row=row_num,column =9, value= total_weight)
        new_sheet.cell(row=row_num,column =14, value=total_vol)
        new_sheet.cell(row=row_num,column =8, value=total_quantity)

        for row in range(8, row_num):
            new_sheet.row_dimensions[row].height = row_height

        align = Alignment(horizontal='center', vertical='center')

        for row in new_sheet.iter_rows(min_row=8, min_col=1, max_col=18, max_row=row_num-1, values_only=False):
            for cell in row:
                cell.border = cell_border
                cell.alignment = align

        output_path = os.path.join(os.getcwd(), 'output')
        today = datetime.date.today()
        formatted_date = today.strftime("%Y.%m.%d")
        output_name = '百泰'+ticket_name[1]+'-'+formatted_date+creat_output_name(ticket_name[2])+'-发票装箱单.xls'
        result_path = os.path.join(output_path, output_name)
        new_ticket.save(result_path)
        os.remove(temp_path)


''' 处理xls文件，将xls文件转为xlsx文件处理'''

def convert_xls_to_xlsx(path):
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False

        for file_name in os.listdir(path):
            if file_name.endswith('.xls'):
                xls_file_path = os.path.join(path, file_name)
                xlsx_file_path = os.path.join(path, file_name[:-4] + '.xlsx')
                workbook = excel.Workbooks.Open(xls_file_path)
                workbook.SaveAs(xlsx_file_path, FileFormat=51)  # FileFormat 51 表示 .xlsx
                workbook.Close()
                os.remove(xls_file_path)  # 删除xls文件
        excel.Quit()
    except Exception as e:
        log.error("：xls文件装换有误联系技术人员检查{}".format(e))


'''解析文件信息'''

# 加载账户信息
def load_account():
    for row in sheet_account.iter_rows(min_row=2, values_only=True):
        # 遍历第6列和第7列
        for cell in row[5:7]:
            account.append(cell)
    if len(account)<1:
        log.error("账户信息表格加载失败")



# 根据文件夹名称记录相关信息,
# 放回发票号和渠道方式
def parse_filename(filename):
    ticket_number =None
    content  = None
    model_name = None

    if filename == None:
        log.error("传入的文件夹为空")
    filename_without_extension = os.path.splitext(filename)[0]      #除去拓展名
    part = filename_without_extension.split("-")
    # 这里分很多种情况，判断前一个或者两个符不符合命名条件这些

    # if filename.startswith("装箱模板 "):
    #     content = filename[len("装箱模板 "):]    #去除前缀
    # else:
    #     content = filename
    #
    # if content ==None:
    #     log.error("文件名为空，请检查文件名称")

    mid_account = part[0] + '-' + part[1]
    max_account = part[0] + '-' + part[1] + '-' +part[2]

    if len(part) < 2:
        # 处理 part 列表不至少有两个元素的情况
        log.error("无效的文件名格式，请检查格式是装箱模板 业务账号（例如：CDMS-发票号(FBA*****)....")
        return None

    if max_account in account:
        ticket_number = part[3]
        model_name = part[4]
    elif mid_account in account:
        ticket_number = part[2]
        model_name = part[3]
    elif part[0] in account:       #这一步直接根据不同情况获取到发票号
        ticket_number = part[1]
        model_name = part[2]
    else:
        log.error("业务账号有误，请检查这个业务号是不是不在相关信息中账号信息的表格中")

    if len(model_name)<3:
        log.error("文件格式有问题")
    else:
        model_name  = model_name[:-3]

    if ticket_number != None and model_name != None:
        return [ticket_number,model_name,filename]
    else:
        log.error("没有找到对应发票号和渠道")
        return None


lindao_model = '林道海运.xlsx'
shunfen_model = '顺丰海运.xlsx'
yinuoda_model = '依诺达.xlsx'
dixing_model = '递信.xlsx'
hongchaung_model = '宏川.xlsx'
e_UPS  =  '欧UPS红单.xlsx'
a_UPS  = '美UPS红单.xlsx'
# 这里要改为根据渠道的名字，之前是根据固定的发票号来的匹配对应的模板的
# 根据发票号找到对的模板
def find_model(name_shipping,input_name):
    # 这里要加一个日志
    xlsx_files = []
    European_contoury = ['英国','意大利','法国','德国','西班牙']
    American_contouty = ['美国','加拿大']

    for f in os.listdir(relate_folder):
        if f.endswith(".xlsx"):
            xlsx_files.append(f)

    if '限时达' in name_shipping:
        name_shipping = name_shipping.replace("限时达",'')

    if 'UPS红单' in name_shipping and '林道' not in input_name[0]:
        for country in European_contoury:
            if country in input_name[0]:
                return e_UPS
        else:
            for country in American_contouty:
                if country in input_name[0]:
                    return a_UPS


    for file_name in xlsx_files:
        if name_shipping in file_name:
            return file_name



    if '林道' in name_shipping:
        return lindao_model
    if '顺丰' in name_shipping:
        return shunfen_model
    if '依诺达' in name_shipping:
        return yinuoda_model
    if '递信' in name_shipping:
        return dixing_model
    if '宏川' in name_shipping:
        return hongchaung_model

    print("检查对应的模板文件是否在模板文件中")
    return None

#读取表格中的SKU，并加入到相关的容器中
# def read_sku(sheet_box_model,index)

# 读取Excel文件，获取商品信息 sheet_box_model是信息所在的sheet,product_sheet是后面要传入的产品导出的表
def read_data_rugular(sheet_box_model, product_sheet):
    box_col = {}
    box_number_index = None

    # if '装箱清单' in sheet_box_model.cell(row=1, column=1).value:
    #找到箱子长宽高和重量信息的位置

    #     for index, row in enumerate(sheet_box_model.iter_cols(min_row=3, max_row=3, values_only=True)):
    #         if '发货数量' == row[0]:
    #             num_index = index + 2
    #             for cell_tuple in sheet_box_model.iter_rows(min_col=num_index):
    #                 for num_cell in cell_tuple:
    #                     if num_cell.value is not None:
    #                         numbers = re.findall(r'\d+', num_cell.value)
    #                         if len(numbers) > 0:  # 只处理包含数字的单元格
    #                             real_num = int(numbers[0])  #转换为int
    #                             # 如果箱号在box_dict中
    #                             if real_num not in box_dict:
    #                                 # 获取箱子对象
    #                                 box_dict[real_num] = []
    #                                 box_real = box_translation(real_num, box_rule_dict[i.offset(row=-1).value],
    #                                                            i.offset(row=-2).value, None)
    #                                 box_dict[real_num] = box_real
    #
    #
    # else:
    # 遍历第一列中的所有单元格
    for index, row in enumerate(sheet_box_model.iter_rows(min_col=2, max_col=2, values_only=True)):
        # 如果单元格的值为“箱号”，则将其索引保存到box_number_index变量中
        if row[0] == '箱号':
            box_number_index = index
            start = index + 1
        # 找到箱号所在行
        if box_number_index is not None:
            # 遍历箱号这一行后面的单元格
            for cell in sheet_box_model.iter_cols(min_row=start, min_col=3):
                # 如果单元格不为空
                for i in cell:
                    if i.row != start:
                        break
                    else:
                        if i.value:
                            # 获取箱号
                            box_num = i.value
                            # 记录箱号的对应行
                            box_col[box_num] = i.column
                            # 如果箱号在box_dict中
                            if box_num not in box_dict:
                                # 获取箱子对象
                                box_dict[box_num] = []
                                box_real = box_translation(box_num, box_rule_dict[i.offset(row=-1).value],
                                                           i.offset(row=-2).value, None)
                                box_dict[box_num] = box_real
                            # 遍历当前单元格所在列
                            for index in sheet_box_model.iter_cols(min_row=start + 1, min_col=box_col[box_num],
                                                                   max_col=box_col[box_num]):
                                # 如果单元格的值不为空，则打印出来
                                for a in index:
                                    if a.value is not None and a.value != '':
                                        box_dict[box_num].array.append(sheet_box_model.cell(row=a.row, column=1).value)

            break

    # 扫sku
    sku_data = []
    sku_index = box_number_index + 1
    for cell in sheet_box_model.iter_rows(min_row=sku_index, min_col=1, max_col=1, values_only=True):
        if cell[0]:
            sku_data.append(cell[0])

    for row in product_sheet.iter_rows(min_row=2, values_only=True):
        sku = row[0]
        if sku in sku_data:
            product_dict[sku] = Product(*row, quantity=None,box_quantity=None)

    # 加入sku的数量属性
    for i, row in enumerate(sheet_box_model.iter_rows(min_row=sku_index + 1, min_col=1, max_col=1, values_only=True),
                            start=sku_index + 1):
        for cell in row:
            product = product_dict.get(cell)
            if product:
                quantity_cell = sheet_box_model.cell(row=i, column=2)
                if quantity_cell.value is not None:
                    product.quantity = quantity_cell.value

    for col in sheet_box_model.iter_cols(min_row=sku_index + 1, min_col=min(box_col.values()),
                                     max_col=max(box_col.values())):
        for cell in col:
            box_quantity = {}  # 存放的是关于箱号和数量
            if cell.value is not None and cell.value != '':
                product_temp = product_dict.get(sheet_box_model.cell(row=cell.row, column=1).value)
                if product_temp is not None:
                    box_n = next((key for key, value in box_col.items() if value == cell.column), None)
                    if box_n is not None:
                        if product_temp.box_quantity is not None and len(product_temp.box_quantity) > 0:
                            product_temp.box_quantity[box_n] = cell.value
                        else:
                            box_quantity[box_n] = cell.value
                            product_temp.box_quantity = box_quantity


'''主程序'''
def process_files():
    # 检查output中是否存在文件，有的话放入history文件夹中
    file_list = os.listdir(output_directory)

    # 检查output文件夹中是否存在文件
    if file_list:
        # 遍历文件列表，将文件转移到history文件夹中
        for file in file_list:
            file_path = os.path.join(output_directory, file)
            if os.path.isfile(file_path):
                history_file_path = os.path.join(history_directory, file)

                # 检查history文件夹中是否存在同名文件
                index = 1
                while os.path.exists(history_file_path):
                    # 文件名已存在，加上序号
                    file_name, file_extension = os.path.splitext(file)
                    new_file_name = f"{file_name}_{index}{file_extension}"
                    history_file_path = os.path.join(history_directory, new_file_name)
                    index += 1

                # 转移文件
                os.rename(file_path, history_file_path)


    '''
        根据文件名中的发货渠道去对应到相应的发票模板，作为一组对应的key，value放入process_file字典中
        接着遍历字典，按不同发票模板的类填入数据
    '''

    #convert_xls_to_xlsx(relate_folder)    #将文件中XLs格式变为XLSx,便于openyxl操作
    load_account()                        #加载相关信息中的账户等相关信息

    input_folder = os.path.join(os.getcwd(), "input")         #定义输入文件夹的位置
    process_files = {}                                        #定义字典用来装要操作的装箱模板

    name_folder = [folder for folder in os.listdir(input_folder) if os.path.isdir(os.path.join(input_folder, folder))]

    processed_folders = [parse_filename(folder) for folder in name_folder]


    #遍历processed_folders,并找到对应模板文件，找将模板文件和装箱单放到process_file的字典中
    #file的内容例子['FBA1764DXC6J', '林道限时达', 'AC-US-FBA1764DXC6J-林道限时达2023.5.19DF票-3件2023.05.19-fba美国计划']
    for file in processed_folders:
        path_tail =os.path.join(input_folder,file[2])
        for data_file in os.listdir(path_tail):
            if '装箱模板' in data_file:
                process_files[os.path.join(path_tail,data_file)] = [find_model(file[1],name_folder),file[0]]

    if len(processed_folders) > len(process_files):
        log.error("输入文件夹中可能缺少装箱模板文件请检查")


    #遍历字典，输出模板
    for file_name, template_name in process_files.items():
        # 因为设定为全局变量，而每一个模板对应box和produc数据都不同，所有每次遍历最开始要清空
        product_dict.clear()
        box_dict.clear()

        # 打开装箱信息文件
        wb = openpyxl.load_workbook(file_name)
        sheet_box_model = wb.active

        # 读取数据
        read_data_rugular(sheet_box_model, sheet_product)
        factory = InvoiceTemplateFactory()
        template_file = os.path.join(relate_folder, template_name[0])

        template_name.append(name_folder)

        if '林道海运' in template_file:
            # try:
                invoice_template = factory.create_template("A")                        #找到对应的模板方法
                invoice_template.fill_data(box_dict, product_dict, template_file,template_name)      #往对应模板方法中传入对应参数
            # except Exception as e:
            #     log.error("模板林道海运带电带磁QT票处出了问题"+ str(e))
        elif '叮铛卡航限时达' in template_file:
            # try:
                invoice_template_b = factory.create_template("B")
                invoice_template_b.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e:
            #     log.error("叮铛卡航限时达QN票模板出了问题"+str(e))
        elif '顺丰海运' in template_file:
            # try:
                invoice_template_c = factory.create_template("C")
                invoice_template_c.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("顺丰空派QA票模板出了问题"+str(e))
        elif '叮铛卡铁' in template_file:
            try:
                invoice_template_d = factory.create_template("D")
                invoice_template_d.fill_data(box_dict, product_dict, template_file,template_name)
            except Exception as e :
                log.error("叮铛卡铁带磁带电QB票出了问题"+str(e))
        elif '叮铛海运' in template_file:
            # try:
                invoice_template_e = factory.create_template("E")
                invoice_template_e.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("叮铛海运带磁带电QL票出了问题"+str(e))
        elif '叮铛美森' in template_file:
            try:
                invoice_template_f = factory.create_template("F")
                invoice_template_f.fill_data(box_dict, product_dict, template_file,template_name)
            except Exception as e :
                log.error("叮铛美森模板出问题"+str(e))
        elif '德邦美森' in template_file:
            # try:
                invoice_template_g = factory.create_template("G")
                invoice_template_g.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("FBA17116ZL52模板出问题"+str(e))
        elif '林道UPS红单' in template_file:
            # try:
                invoice_template_i = factory.create_template("I")
                invoice_template_i.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("林道红单模板出问题"+str(e))
        elif '美UPS红单' in template_file:
            # try:
                invoice_template_j = factory.create_template("J")
                invoice_template_j.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("红单DJ票"+str(e))
        elif '德邦空派' in template_file:
            # try:
                invoice_template_k = factory.create_template("K")
                invoice_template_k.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("FBA171F1JKW8模板出问题"+str(e))
        elif '欧UPS红单' in template_file:
            try:
                invoice_template_L = factory.create_template("L")
                invoice_template_L.fill_data(box_dict, product_dict, template_file,template_name)
            except Exception as e :
                log.error("FBA15GMRY8YG模板出问题"+str(e))
        elif '递信' in template_file:
            # try:
                invoice_template_M = factory.create_template("M")
                invoice_template_M.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("FBA15DCLNMJP模板出问题"+str(e))
        elif '宏川' in template_file:
            # try:
                invoice_template_N = factory.create_template("N")
                invoice_template_N.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("宏川"+str(e))
        elif '依诺达' in template_file:
            # try:
                invoice_template_H = factory.create_template("H")
                invoice_template_H.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("H模板"+str(e))
        elif '林道日本' in template_file:
            # try:
                invoice_template_H = factory.create_template("O")
                invoice_template_H.fill_data(box_dict, product_dict, template_file,template_name)
            # except Exception as e :
            #     log.error("H模板"+str(e))
        else:
            log.error("没有找到对应的模板文件")



#检查远程仓库是否有变动
subprocess.run(['git','fetch'])

diff_output  = subprocess.run(['git','differ','origin/master'],capture_output=True,text=True).stdout


#主函数
if __name__ == "__main__":
    log = mylogger("test", console_print=True, level=20, log_file=log_file)

    try:
        process_files()
    except Exception as e:
        log.error(traceback.format_exc())
    #检查代码是否有变动
    if diff_output:
        root = tk.TK()
        root.witdraw()
        messagebox.showinfo("有新版本出现，请及时更新")
        root.destory()
